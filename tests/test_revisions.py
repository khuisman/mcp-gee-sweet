import io

import openpyxl

from mcp_gee_sweet.tools.drive import _xlsx_range_values


def _make_wb(data: list[list]) -> openpyxl.Workbook:
    """Build an in-memory workbook with data written to Sheet1."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in data:
        ws.append(row)
    return wb


def _roundtrip(wb: openpyxl.Workbook) -> openpyxl.Workbook:
    """Save to bytes and reload read-only (mirrors what export_revision does)."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return openpyxl.load_workbook(buf, read_only=True, data_only=True)


class TestXlsxRangeValues:
    def test_no_range_returns_all_rows(self):
        wb = _roundtrip(_make_wb([["A", "B"], ["C", "D"]]))
        result = _xlsx_range_values(wb.active, None)
        assert result == [["A", "B"], ["C", "D"]]

    def test_multi_cell_range(self):
        wb = _roundtrip(_make_wb([["A", "B", "C"], ["D", "E", "F"], ["G", "H", "I"]]))
        result = _xlsx_range_values(wb.active, "A1:B2")
        assert result == [["A", "B"], ["D", "E"]]

    def test_single_row_range(self):
        wb = _roundtrip(_make_wb([["X", "Y", "Z"]]))
        result = _xlsx_range_values(wb.active, "A1:C1")
        assert result == [["X", "Y", "Z"]]

    def test_single_cell_range(self):
        wb = _roundtrip(_make_wb([["Hello", "World"]]))
        result = _xlsx_range_values(wb.active, "A1")
        assert result == [["Hello"]]

    def test_empty_cells_return_none(self):
        wb = _roundtrip(_make_wb([["A", None, "C"]]))
        result = _xlsx_range_values(wb.active, "A1:C1")
        assert result == [["A", None, "C"]]

    def test_numeric_values(self):
        wb = _roundtrip(_make_wb([[1, 2.5, 3]]))
        result = _xlsx_range_values(wb.active, "A1:C1")
        assert result == [[1, 2.5, 3]]

    def test_second_sheet(self):
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["only", "here"])
        rb = _roundtrip(wb)
        result = _xlsx_range_values(rb["Sheet2"], "A1:B1")
        assert result == [["only", "here"]]
