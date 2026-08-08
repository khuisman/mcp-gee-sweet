"""Tests for tools/drive/transfer.py (upload_file, _xlsx_range_values, etc.)."""

import io
import json
import os
import threading
from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock

import openpyxl
import pytest
from googleapiclient.errors import HttpError

from mcp_gee_sweet.tools import response_limits
from mcp_gee_sweet.tools.drive import transfer as transfer_module
from mcp_gee_sweet.tools.drive.transfer import _upload_local_file, _xlsx_range_values


def _make_tool_registry():
    captured = {}

    def tool(annotations=None):
        def decorator(func):
            captured[func.__name__] = func
            return func

        return decorator

    return tool, captured


def _make_ctx(**services):
    ctx = MagicMock()
    lc = ctx.request_context.lifespan_context
    for k, v in services.items():
        setattr(lc, k, v)
    return ctx


_transfer_tool, _transfer_tools = _make_tool_registry()
transfer_module.register(_transfer_tool)


def _make_wb(data: list[list]) -> openpyxl.Workbook:
    """Build an in-memory workbook with data written to Sheet1."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in data:
        ws.append(row)
    return wb


def _roundtrip(wb: openpyxl.Workbook) -> openpyxl.Workbook:
    """Save to bytes and reload read-only (mirrors what export_revision does)."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return openpyxl.load_workbook(buf, read_only=True, data_only=True)


class TestUploadFile:
    """upload_file returns a friendly error on quota exceeded and invalidates the folder cache."""

    def _quota_err(self):
        resp = MagicMock()
        resp.status = 403
        return HttpError(resp=resp, content=b'{"error": {"reason": "storageQuotaExceeded"}}')

    def _drive_file_response(self):
        return {
            "id": "fid1",
            "name": "file.txt",
            "parents": ["parent1"],
            "mimeType": "text/plain",
            "webViewLink": "https://example.com",
        }

    async def test_quota_exceeded_returns_friendly_error_dict(self):
        """upload_file must return {"error": ...} on storageQuotaExceeded, not raise."""
        mock = MagicMock()
        mock.files.return_value.create.return_value.execute.side_effect = self._quota_err()
        ctx = _make_ctx(drive_service=mock, drive_folder_cache=MagicMock(), folder_id=None)
        result = await _transfer_tools["upload_file"](name="test.txt", content="hello", ctx=ctx)
        assert "error" in result
        assert "storageQuotaExceeded" not in result["error"]  # raw message replaced
        assert "Service accounts" in result["error"]
        assert "server://auth-status" in result["error"]

    async def test_with_folder_marks_folder_cache_dirty(self):
        mock = MagicMock()
        mock.files.return_value.create.return_value.execute.return_value = (
            self._drive_file_response()
        )
        folder_cache = MagicMock()
        ctx = _make_ctx(
            drive_service=mock, drive_folder_cache=folder_cache, folder_id="default_folder"
        )
        await _transfer_tools["upload_file"](
            name="doc.txt", content="hello", folder_id="target_folder", ctx=ctx
        )
        folder_cache.mark_dirty.assert_called_once_with("target_folder")


class TestUploadLocalFileCore:
    """Direct tests for _upload_local_file — the module-level helper factored out
    of the upload_local_file tool so docs/content.py's insert_local_images can
    call it directly."""

    def _quota_err(self):
        resp = MagicMock()
        resp.status = 403
        return HttpError(resp=resp, content=b'{"error": {"reason": "storageQuotaExceeded"}}')

    async def test_missing_local_file_raises(self, tmp_path):
        drive_svc = MagicMock()
        with pytest.raises(ValueError, match="No file found"):
            await _upload_local_file(drive_svc, str(tmp_path / "missing.txt"), "folder1")

    async def test_uploads_and_returns_file_metadata(self, tmp_path):
        local_file = tmp_path / "pic.png"
        local_file.write_bytes(b"fake-bytes")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {"files": []}
        drive_svc.files.return_value.create.return_value.execute.return_value = {
            "id": "fid1",
            "name": "pic.png",
            "webViewLink": "https://example.com/pic",
        }

        result = await _upload_local_file(drive_svc, str(local_file), "folder1")

        assert result == {
            "fileId": "fid1",
            "name": "pic.png",
            "web_link": "https://example.com/pic",
            "skipped": False,
        }
        create_kwargs = drive_svc.files.return_value.create.call_args.kwargs
        body = create_kwargs["body"]
        assert body["name"] == "pic.png"
        assert body["parents"] == ["folder1"]
        # #274 PR #472 review, finding #2: a plain (non-convert) upload now stamps
        # modifiedTime from the local file's mtime too, not just the convert branch.
        assert "modifiedTime" in body

    async def test_skip_if_exists_returns_existing_file_without_uploading(self, tmp_path):
        local_file = tmp_path / "pic.png"
        local_file.write_bytes(b"fake-bytes")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {
            "files": [{"id": "existing1", "name": "pic.png", "webViewLink": "https://x/existing"}]
        }

        result = await _upload_local_file(
            drive_svc, str(local_file), "folder1", skip_if_exists=True
        )

        assert result == {
            "fileId": "existing1",
            "name": "pic.png",
            "web_link": "https://x/existing",
            "skipped": True,
        }
        drive_svc.files.return_value.create.assert_not_called()

    async def test_quota_exceeded_returns_friendly_error_dict(self, tmp_path):
        local_file = tmp_path / "pic.png"
        local_file.write_bytes(b"fake-bytes")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {"files": []}
        drive_svc.files.return_value.create.return_value.execute.side_effect = self._quota_err()

        result = await _upload_local_file(drive_svc, str(local_file), "folder1")

        assert "error" in result
        assert "Service accounts" in result["error"]

    async def test_custom_name_used_instead_of_filename(self, tmp_path):
        local_file = tmp_path / "pic.png"
        local_file.write_bytes(b"fake-bytes")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {"files": []}
        drive_svc.files.return_value.create.return_value.execute.return_value = {
            "id": "fid1",
            "name": "renamed.png",
            "webViewLink": "https://example.com/pic",
        }

        result = await _upload_local_file(drive_svc, str(local_file), "folder1", name="renamed.png")

        assert result["name"] == "renamed.png"
        create_kwargs = drive_svc.files.return_value.create.call_args.kwargs
        assert create_kwargs["body"]["name"] == "renamed.png"


class TestUploadLocalFileConvert:
    """convert=True requests Drive's native import conversion (issue #188)."""

    @pytest.mark.parametrize(
        "filename,expected_target_mime",
        [
            ("data.csv", "application/vnd.google-apps.spreadsheet"),
            ("data.xlsx", "application/vnd.google-apps.spreadsheet"),
            ("doc.docx", "application/vnd.google-apps.document"),
            ("notes.md", "application/vnd.google-apps.document"),
            ("page.html", "application/vnd.google-apps.document"),
            ("page.htm", "application/vnd.google-apps.document"),
            ("deck.pptx", "application/vnd.google-apps.presentation"),
        ],
    )
    async def test_convert_sets_target_mimetype_for_supported_extensions(
        self, tmp_path, filename, expected_target_mime
    ):
        local_file = tmp_path / filename
        local_file.write_text("content")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {"files": []}
        drive_svc.files.return_value.create.return_value.execute.return_value = {
            "id": "fid1",
            "name": filename,
            "webViewLink": "https://example.com",
        }

        result = await _upload_local_file(
            drive_svc, str(local_file), "folder1", skip_if_exists=False, convert=True
        )

        assert "error" not in result
        create_kwargs = drive_svc.files.return_value.create.call_args.kwargs
        assert create_kwargs["body"]["mimeType"] == expected_target_mime

    async def test_convert_md_stamps_source_property_for_sync_folder_matching(self, tmp_path):
        """#414 QA review round 3, finding #2: upload_local_file(convert=True) never
        stamped the marker sync_folder's own convert_markdown matching relies on,
        so a Doc created here was invisible to that matching and got silently
        duplicated on the next sync_folder run. Only .md needs this — it's the only
        extension sync_folder's convert_markdown has an equivalent matching path
        for."""
        local_file = tmp_path / "notes.md"
        local_file.write_text("# Heading")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {"files": []}
        drive_svc.files.return_value.create.return_value.execute.return_value = {
            "id": "fid1",
            "name": "notes.md",
            "webViewLink": "https://example.com",
        }

        result = await _upload_local_file(
            drive_svc, str(local_file), "folder1", skip_if_exists=False, convert=True
        )

        assert "error" not in result
        create_kwargs = drive_svc.files.return_value.create.call_args.kwargs
        assert create_kwargs["body"]["properties"] == {
            transfer_module._CONVERT_MARKDOWN_SOURCE_PROP: "notes.md"
        }

    async def test_convert_stamps_modified_time_from_local_mtime_and_restamps_after_create(
        self, tmp_path
    ):
        """#422 finding #2: _upload_local_file never set modifiedTime on a converted
        Doc, so it got Drive's own creation timestamp instead of the local file's
        mtime — landing in sync_folder's 'conflicts' rather than 'skipped' on
        nearly every first sync_folder call afterward, since nothing tied the two
        timestamps together. A metadata-only follow-up update() re-stamps it after
        Drive's native import overwrites the create() request, the same way
        _sync_level's own convert_markdown upload path already does (TC-D218)."""
        local_file = tmp_path / "notes.md"
        local_file.write_text("# Heading")
        expected_mtime = datetime.fromtimestamp(
            local_file.stat().st_mtime, tz=timezone.utc
        ).strftime("%Y-%m-%dT%H:%M:%S.000Z")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {"files": []}
        drive_svc.files.return_value.create.return_value.execute.return_value = {
            "id": "fid1",
            "name": "notes.md",
            "webViewLink": "https://example.com",
        }
        drive_svc.files.return_value.update.return_value.execute.return_value = {"id": "fid1"}

        result = await _upload_local_file(
            drive_svc, str(local_file), "folder1", skip_if_exists=False, convert=True
        )

        assert "error" not in result
        create_kwargs = drive_svc.files.return_value.create.call_args.kwargs
        assert create_kwargs["body"]["modifiedTime"] == expected_mtime

        update_kwargs = drive_svc.files.return_value.update.call_args.kwargs
        assert update_kwargs["fileId"] == "fid1"
        assert update_kwargs["body"] == {"modifiedTime": expected_mtime}
        assert "media_body" not in update_kwargs

    async def test_convert_modified_time_restamp_failure_returns_clean_error_not_raise(
        self, tmp_path
    ):
        """#422 finding #1: a failure in the follow-up modifiedTime-restamp update()
        call — after create() already succeeded — used to propagate as an
        uncaught exception, even though a Doc now genuinely exists in Drive. The
        caller (the upload_local_file tool) has no try/except of its own, so this
        would otherwise surface as an unhandled tool error with no fileId
        returned at all. Mirrors _sync_level._run_one's create()+update() pair,
        which already wraps both calls in one broad except and returns a clean
        failure rather than letting a partial success propagate raw."""
        local_file = tmp_path / "notes.md"
        local_file.write_text("# Heading")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {"files": []}
        drive_svc.files.return_value.create.return_value.execute.return_value = {
            "id": "fid1",
            "name": "notes.md",
            "webViewLink": "https://example.com",
        }
        drive_svc.files.return_value.update.return_value.execute.side_effect = RuntimeError(
            "transient network error"
        )

        result = await _upload_local_file(
            drive_svc, str(local_file), "folder1", skip_if_exists=False, convert=True
        )

        assert "error" in result
        assert "transient network error" in result["error"]

    async def test_convert_non_md_extension_does_not_stamp_source_property(self, tmp_path):
        local_file = tmp_path / "data.csv"
        local_file.write_text("a,b\n1,2")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {"files": []}
        drive_svc.files.return_value.create.return_value.execute.return_value = {
            "id": "fid1",
            "name": "data.csv",
            "webViewLink": "https://example.com",
        }

        result = await _upload_local_file(
            drive_svc, str(local_file), "folder1", skip_if_exists=False, convert=True
        )

        assert "error" not in result
        create_kwargs = drive_svc.files.return_value.create.call_args.kwargs
        assert "properties" not in create_kwargs["body"]

    async def test_convert_unsupported_extension_returns_error_without_uploading(self, tmp_path):
        local_file = tmp_path / "archive.zip"
        local_file.write_bytes(b"fake-bytes")
        drive_svc = MagicMock()

        result = await _upload_local_file(
            drive_svc, str(local_file), "folder1", skip_if_exists=False, convert=True
        )

        assert "error" in result
        assert ".zip" in result["error"]
        drive_svc.files.return_value.create.assert_not_called()

    async def test_convert_false_default_does_not_set_target_mimetype(self, tmp_path):
        local_file = tmp_path / "data.csv"
        local_file.write_text("a,b\n1,2")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {"files": []}
        drive_svc.files.return_value.create.return_value.execute.return_value = {
            "id": "fid1",
            "name": "data.csv",
            "webViewLink": "https://example.com",
        }

        result = await _upload_local_file(drive_svc, str(local_file), "folder1")

        assert "error" not in result
        create_kwargs = drive_svc.files.return_value.create.call_args.kwargs
        assert "mimeType" not in create_kwargs["body"]
        # modifiedTime is stamped on every upload now, not just the convert branch
        # (#274 PR #472 review, finding #2) — but only via the create() body itself;
        # no follow-up restamp is needed (or performed) for a plain upload, since
        # Drive has no import-conversion here to overwrite the value.
        assert "modifiedTime" in create_kwargs["body"]
        drive_svc.files.return_value.update.assert_not_called()

    async def test_convert_extension_comes_from_name_override_not_local_path(self, tmp_path):
        """A no-extension local_path with a .csv name= override should still convert
        (PR #410 QA review: the lookup used local_path's suffix, not the effective
        destination name, so this previously errored as "unsupported extension")."""
        local_file = tmp_path / "scratch_tmpfile"
        local_file.write_text("a,b\n1,2")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {"files": []}
        drive_svc.files.return_value.create.return_value.execute.return_value = {
            "id": "fid1",
            "name": "report.csv",
            "webViewLink": "https://example.com",
        }

        result = await _upload_local_file(
            drive_svc,
            str(local_file),
            "folder1",
            name="report.csv",
            skip_if_exists=False,
            convert=True,
        )

        assert "error" not in result
        create_kwargs = drive_svc.files.return_value.create.call_args.kwargs
        assert create_kwargs["body"]["mimeType"] == "application/vnd.google-apps.spreadsheet"

    async def test_convert_extension_mismatch_between_local_path_and_name_override_errors(
        self, tmp_path
    ):
        """Inverse of the above: a .csv local_path with a name= override that has no
        supported extension should error on the destination extension, not silently
        succeed using local_path's .csv (PR #410 QA review)."""
        local_file = tmp_path / "upload.csv"
        local_file.write_text("a,b\n1,2")
        drive_svc = MagicMock()

        result = await _upload_local_file(
            drive_svc,
            str(local_file),
            "folder1",
            name="archive.zip",
            skip_if_exists=False,
            convert=True,
        )

        assert "error" in result
        assert ".zip" in result["error"]
        drive_svc.files.return_value.create.assert_not_called()

    async def test_skip_if_exists_does_not_skip_when_existing_file_is_unconverted(self, tmp_path):
        """skip_if_exists must not treat a same-named raw (unconverted) file as the
        skip-worthy duplicate when convert=True — a name-only match previously
        returned the raw file with no conversion and no error (PR #410 QA review)."""
        local_file = tmp_path / "a.csv"
        local_file.write_text("a,b\n1,2")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {
            "files": [
                {
                    "id": "existing-raw",
                    "name": "a.csv",
                    "webViewLink": "https://x/existing",
                    "mimeType": "text/csv",
                }
            ]
        }
        drive_svc.files.return_value.create.return_value.execute.return_value = {
            "id": "fid-converted",
            "name": "a.csv",
            "webViewLink": "https://example.com/converted",
        }

        result = await _upload_local_file(drive_svc, str(local_file), "folder1", convert=True)

        assert result["skipped"] is False
        assert result["fileId"] == "fid-converted"
        create_kwargs = drive_svc.files.return_value.create.call_args.kwargs
        assert create_kwargs["body"]["mimeType"] == "application/vnd.google-apps.spreadsheet"

    async def test_skip_if_exists_still_skips_when_existing_file_already_converted(self, tmp_path):
        """The converted case that skip_if_exists is actually meant to catch: an
        existing file already in the target Workspace mimeType should still skip."""
        local_file = tmp_path / "a.csv"
        local_file.write_text("a,b\n1,2")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {
            "files": [
                {
                    "id": "existing-converted",
                    "name": "a.csv",
                    "webViewLink": "https://x/existing",
                    "mimeType": "application/vnd.google-apps.spreadsheet",
                }
            ]
        }

        result = await _upload_local_file(drive_svc, str(local_file), "folder1", convert=True)

        assert result == {
            "fileId": "existing-converted",
            "name": "a.csv",
            "web_link": "https://x/existing",
            "skipped": True,
        }
        drive_svc.files.return_value.create.assert_not_called()


class TestUploadLocalFolder:
    """upload_local_folder now routes each file through the shared
    _upload_local_file helper (issue #411) instead of its own independent
    inline create() call, so it can offer the same convert param
    upload_local_file already has. The bulk skip_if_exists pre-check (one
    list() call for the whole folder, TC-D100) is preserved rather than
    delegated per-file, since _upload_local_file's own skip_if_exists check
    would cost one list() call per candidate instead."""

    def _tool(self):
        return _transfer_tools["upload_local_folder"]

    def _ctx(self, drive_svc):
        return _make_ctx(drive_service=drive_svc, drive_folder_cache=MagicMock())

    async def test_bulk_upload_of_mixed_directory(self, tmp_path):
        (tmp_path / "a.txt").write_text("a")
        (tmp_path / "b.png").write_bytes(b"fake")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {"files": []}
        drive_svc.files.return_value.create.return_value.execute.return_value = {
            "id": "fid",
            "name": "x",
            "webViewLink": "https://example.com",
        }
        ctx = self._ctx(drive_svc)

        result = await self._tool()(str(tmp_path), "folder1", ctx=ctx)

        assert result["uploaded"] == ["a.txt", "b.png"]
        assert result["skipped"] == []
        assert result["failed"] == []
        ctx.request_context.lifespan_context.drive_folder_cache.mark_dirty.assert_called_once_with(
            "folder1"
        )

    async def test_ds_store_excluded_by_default(self, tmp_path):
        (tmp_path / "a.txt").write_text("a")
        (tmp_path / ".DS_Store").write_bytes(b"junk")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {"files": []}
        drive_svc.files.return_value.create.return_value.execute.return_value = {
            "id": "fid",
            "name": "a.txt",
            "webViewLink": "https://example.com",
        }

        result = await self._tool()(str(tmp_path), "folder1", ctx=self._ctx(drive_svc))

        assert result["uploaded"] == ["a.txt"]
        assert ".DS_Store" not in result["uploaded"]

    async def test_skip_if_exists_makes_exactly_one_list_call_for_the_whole_folder(self, tmp_path):
        """TC-D100: the existence check must stay a single batched list() call
        across the whole folder, not one per file — this is why the bulk
        pre-check result is passed down as skip_if_exists=False into
        _upload_local_file rather than letting each call do its own check."""
        (tmp_path / "a.txt").write_text("a")
        (tmp_path / "b.txt").write_text("b")
        (tmp_path / "c.txt").write_text("c")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {
            "files": [{"name": "a.txt", "mimeType": "text/plain"}]
        }
        drive_svc.files.return_value.create.return_value.execute.return_value = {
            "id": "fid",
            "name": "x",
            "webViewLink": "https://example.com",
        }

        result = await self._tool()(str(tmp_path), "folder1", ctx=self._ctx(drive_svc))

        assert result["skipped"] == ["a.txt"]
        assert sorted(result["uploaded"]) == ["b.txt", "c.txt"]
        assert drive_svc.files.return_value.list.call_count == 1

    async def test_convert_routes_through_shared_helper_and_sets_target_mimetype(self, tmp_path):
        (tmp_path / "data.csv").write_text("a,b\n1,2")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {"files": []}
        drive_svc.files.return_value.create.return_value.execute.return_value = {
            "id": "fid",
            "name": "data.csv",
            "webViewLink": "https://example.com",
        }
        drive_svc.files.return_value.update.return_value.execute.return_value = {"id": "fid"}

        result = await self._tool()(
            str(tmp_path), "folder1", convert=True, ctx=self._ctx(drive_svc)
        )

        assert result["uploaded"] == ["data.csv"]
        assert result["failed"] == []
        create_kwargs = drive_svc.files.return_value.create.call_args.kwargs
        assert create_kwargs["body"]["mimeType"] == "application/vnd.google-apps.spreadsheet"

    async def test_convert_unsupported_extension_reported_as_failed_not_uploaded(self, tmp_path):
        (tmp_path / "archive.zip").write_bytes(b"fake")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {"files": []}

        result = await self._tool()(
            str(tmp_path), "folder1", convert=True, ctx=self._ctx(drive_svc)
        )

        assert result["uploaded"] == []
        assert len(result["failed"]) == 1
        assert result["failed"][0]["name"] == "archive.zip"
        assert ".zip" in result["failed"][0]["error"]
        drive_svc.files.return_value.create.assert_not_called()

    async def test_convert_skips_when_existing_file_already_in_target_mimetype(self, tmp_path):
        (tmp_path / "data.csv").write_text("a,b\n1,2")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {
            "files": [{"name": "data.csv", "mimeType": "application/vnd.google-apps.spreadsheet"}]
        }

        result = await self._tool()(
            str(tmp_path), "folder1", convert=True, ctx=self._ctx(drive_svc)
        )

        assert result["skipped"] == ["data.csv"]
        assert result["uploaded"] == []
        drive_svc.files.return_value.create.assert_not_called()

    async def test_convert_does_not_skip_when_existing_file_is_still_unconverted(self, tmp_path):
        """A same-named file that exists but hasn't been converted yet (still its
        raw mimeType) is not the duplicate convert=True's skip check is meant to
        catch — mirrors _upload_local_file's own convert-aware skip logic."""
        (tmp_path / "data.csv").write_text("a,b\n1,2")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {
            "files": [{"name": "data.csv", "mimeType": "text/csv"}]
        }
        drive_svc.files.return_value.create.return_value.execute.return_value = {
            "id": "fid-converted",
            "name": "data.csv",
            "webViewLink": "https://example.com",
        }
        drive_svc.files.return_value.update.return_value.execute.return_value = {
            "id": "fid-converted"
        }

        result = await self._tool()(
            str(tmp_path), "folder1", convert=True, ctx=self._ctx(drive_svc)
        )

        assert result["uploaded"] == ["data.csv"]
        assert result["skipped"] == []

    async def test_convert_skips_when_existing_converted_file_has_stripped_extension(
        self, tmp_path
    ):
        """TC-D243 (PR #505 review): Drive's native import-conversion strips the
        source extension from some converted types' display name (confirmed
        live for CSV) — a second convert=True run must still recognize the
        already-converted file even though its Drive name ("data") no longer
        matches the local file's name ("data.csv"), or it re-converts and
        duplicates it on every run."""
        (tmp_path / "data.csv").write_text("a,b\n1,2")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {
            "files": [{"name": "data", "mimeType": "application/vnd.google-apps.spreadsheet"}]
        }

        result = await self._tool()(
            str(tmp_path), "folder1", convert=True, ctx=self._ctx(drive_svc)
        )

        assert result["skipped"] == ["data.csv"]
        assert result["uploaded"] == []
        drive_svc.files.return_value.create.assert_not_called()

    async def test_convert_skips_when_both_raw_and_converted_duplicates_exist(self, tmp_path):
        """A third convert=True run against a folder that already has both the
        raw pre-convert duplicate (TC-D242's scenario, matched by full name) and
        the already-converted file (TC-D243's scenario, matched by stem) must
        still recognize the converted one and skip, not create a third file —
        the two lookups must be independent, not short-circuit on the first
        (unrelated) match."""
        (tmp_path / "data.csv").write_text("a,b\n1,2")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {
            "files": [
                {"name": "data.csv", "mimeType": "text/csv"},
                {"name": "data", "mimeType": "application/vnd.google-apps.spreadsheet"},
            ]
        }

        result = await self._tool()(
            str(tmp_path), "folder1", convert=True, ctx=self._ctx(drive_svc)
        )

        assert result["skipped"] == ["data.csv"]
        assert result["uploaded"] == []
        drive_svc.files.return_value.create.assert_not_called()

    async def test_file_deleted_mid_loop_reported_as_failed_not_raised(self, tmp_path, monkeypatch):
        """_upload_local_file raises ValueError uncaught if the local file is
        gone by the time it's actually called (its own path.is_file() check
        runs before its try/except) — e.g. deleted between the directory scan
        and this file's turn in the loop. The old inline upload code caught
        this per-file; the refactor to _upload_local_file needs its own
        try/except to keep that isolation (PR #505 review, issue #411).
        Simulated via monkeypatch since a real race on disk isn't reliably
        reproducible in a unit test."""
        (tmp_path / "a.txt").write_text("a")
        (tmp_path / "b.txt").write_text("b")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {"files": []}

        real_upload = transfer_module._upload_local_file

        async def _flaky(drive_service, local_path, *args, **kwargs):
            if local_path.endswith("b.txt"):
                raise ValueError(f"No file found at {local_path!r}")
            return await real_upload(drive_service, local_path, *args, **kwargs)

        monkeypatch.setattr(transfer_module, "_upload_local_file", _flaky)
        drive_svc.files.return_value.create.return_value.execute.return_value = {
            "id": "fid",
            "name": "a.txt",
            "webViewLink": "https://example.com",
        }

        result = await self._tool()(str(tmp_path), "folder1", ctx=self._ctx(drive_svc))

        assert result["uploaded"] == ["a.txt"]
        assert len(result["failed"]) == 1
        assert result["failed"][0]["name"] == "b.txt"
        assert "No file found" in result["failed"][0]["error"]

    async def test_create_failure_reported_as_failed_not_raised(self, tmp_path):
        (tmp_path / "a.txt").write_text("a")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {"files": []}
        drive_svc.files.return_value.create.return_value.execute.side_effect = RuntimeError("boom")

        result = await self._tool()(str(tmp_path), "folder1", ctx=self._ctx(drive_svc))

        assert result["uploaded"] == []
        assert result["failed"] == [{"name": "a.txt", "error": "boom"}]

    async def test_no_upload_does_not_mark_folder_cache_dirty(self, tmp_path):
        (tmp_path / "a.txt").write_text("a")
        drive_svc = MagicMock()
        drive_svc.files.return_value.list.return_value.execute.return_value = {
            "files": [{"name": "a.txt", "mimeType": "text/plain"}]
        }
        ctx = self._ctx(drive_svc)

        result = await self._tool()(str(tmp_path), "folder1", ctx=ctx)

        assert result["uploaded"] == []
        ctx.request_context.lifespan_context.drive_folder_cache.mark_dirty.assert_not_called()

    async def test_missing_local_directory_raises(self, tmp_path):
        drive_svc = MagicMock()
        with pytest.raises(ValueError, match="No directory found"):
            await self._tool()(str(tmp_path / "missing"), "folder1", ctx=self._ctx(drive_svc))


class TestXlsxRangeValues:
    async def test_no_range_returns_all_rows(self):
        wb = _roundtrip(_make_wb([["A", "B"], ["C", "D"]]))
        result = _xlsx_range_values(wb.active, None)
        assert result == [["A", "B"], ["C", "D"]]

    async def test_multi_cell_range(self):
        wb = _roundtrip(_make_wb([["A", "B", "C"], ["D", "E", "F"], ["G", "H", "I"]]))
        result = _xlsx_range_values(wb.active, "A1:B2")
        assert result == [["A", "B"], ["D", "E"]]

    async def test_single_row_range(self):
        wb = _roundtrip(_make_wb([["X", "Y", "Z"]]))
        result = _xlsx_range_values(wb.active, "A1:C1")
        assert result == [["X", "Y", "Z"]]

    async def test_single_cell_range(self):
        wb = _roundtrip(_make_wb([["Hello", "World"]]))
        result = _xlsx_range_values(wb.active, "A1")
        assert result == [["Hello"]]

    async def test_empty_cells_return_none(self):
        wb = _roundtrip(_make_wb([["A", None, "C"]]))
        result = _xlsx_range_values(wb.active, "A1:C1")
        assert result == [["A", None, "C"]]

    async def test_numeric_values(self):
        wb = _roundtrip(_make_wb([[1, 2.5, 3]]))
        result = _xlsx_range_values(wb.active, "A1:C1")
        assert result == [[1, 2.5, 3]]

    async def test_second_sheet(self):
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["only", "here"])
        rb = _roundtrip(wb)
        result = _xlsx_range_values(rb["Sheet2"], "A1:B1")
        assert result == [["only", "here"]]


class TestExportFile:
    """export_file's base64 encoding inflates raw size ~33%, so it needs the response-size
    safety net too (issue #242) — but points to download_file instead of a local_path
    bypass, since download_file already writes raw bytes with no base64/JSON overhead."""

    def _ctx(self, drive_svc):
        ctx = MagicMock()
        ctx.request_context.lifespan_context.drive_service = drive_svc
        return ctx

    def _workspace_svc(self, content_bytes, mime_type="application/vnd.google-apps.document"):
        svc = MagicMock()
        svc.files.return_value.get.return_value.execute.return_value = {
            "id": "doc1",
            "name": "Test Doc",
            "mimeType": mime_type,
        }
        svc.files.return_value.export.return_value.execute.return_value = content_bytes
        return svc

    async def test_small_export_succeeds(self):
        ctx = self._ctx(self._workspace_svc(b"small pdf content"))
        result = await _transfer_tools["export_file"](file_id="doc1", export_format="pdf", ctx=ctx)
        assert result["encoding"] == "base64"

    async def test_oversized_base64_content_raises(self, monkeypatch):
        monkeypatch.setattr(response_limits, "MAX_TOOL_RESPONSE_CHARS", 10)
        ctx = self._ctx(self._workspace_svc(b"x" * 1000))
        with pytest.raises(ValueError, match="safety cap"):
            await _transfer_tools["export_file"](file_id="doc1", export_format="pdf", ctx=ctx)

    async def test_error_points_to_download_file_not_local_path(self, monkeypatch):
        # export_file has no local_path param — download_file is the correct bypass
        # (raw bytes to disk, no base64/JSON overhead), so the error must say so and
        # must not reference a param this tool doesn't have.
        monkeypatch.setattr(response_limits, "MAX_TOOL_RESPONSE_CHARS", 10)
        ctx = self._ctx(self._workspace_svc(b"x" * 1000))
        with pytest.raises(ValueError) as exc_info:
            await _transfer_tools["export_file"](file_id="doc1", export_format="pdf", ctx=ctx)
        msg = str(exc_info.value)
        assert "download_file" in msg
        assert "local_path" not in msg


class _FakeDriveFS:
    """Fakes the slice of the Drive API sync_folder needs: per-folder `files().list`,
    plus `files().create`/`files().update` that record what would have been written."""

    def __init__(self, children: dict[str, list[dict]]):
        self.children = children  # folder_id -> Drive API file/folder dicts
        self.list_calls: list[str] = []  # folder ids queried, in order
        self.created_folders: list[dict] = []
        self.created_files: list[dict] = []
        self.updated_files: list[dict] = []

        self.svc = MagicMock()
        self.svc.files.return_value.list.side_effect = self._list
        self.svc.files.return_value.create.side_effect = self._create
        self.svc.files.return_value.update.side_effect = self._update

    def _list(self, **kwargs):
        folder_id = kwargs["q"].split("'")[1]
        self.list_calls.append(folder_id)
        resp = MagicMock()
        resp.execute.return_value = {"files": self.children.get(folder_id, [])}
        return resp

    def _create(self, **kwargs):
        body = kwargs["body"]
        resp = MagicMock()
        if body.get("mimeType") == "application/vnd.google-apps.folder":
            new_id = f"new-folder-{len(self.created_folders)}"
            self.created_folders.append(body)
            resp.execute.return_value = {"id": new_id}
        else:
            self.created_files.append(body)
            resp.execute.return_value = {"id": f"new-file-{len(self.created_files)}"}
        return resp

    def _update(self, **kwargs):
        self.updated_files.append(kwargs)
        resp = MagicMock()
        resp.execute.return_value = {"id": kwargs.get("fileId")}
        return resp


def _drive_file(
    name,
    file_id,
    mtime="2024-01-01T00:00:00.000Z",
    mime="text/plain",
    properties=None,
    md5=None,
):
    f = {"id": file_id, "name": name, "mimeType": mime, "modifiedTime": mtime}
    if properties is not None:
        f["properties"] = properties
    if md5 is not None:
        f["md5Checksum"] = md5
    return f


def _drive_folder(name, folder_id):
    return {"id": folder_id, "name": name, "mimeType": "application/vnd.google-apps.folder"}


class TestSyncFolderRecursive:
    """Issue #315: sync_folder silently ignored every subfolder, one level deep,
    reporting a clean 'in sync' result instead of surfacing the gap. `recursive=True`
    now walks matching subfolders to any depth; subfolders left alone because the
    sync direction wouldn't create them on the missing side are reported under
    'folders_skipped' instead of vanishing silently."""

    def _ctx(self, fs: _FakeDriveFS):
        ctx = MagicMock()
        ctx.request_context.lifespan_context.drive_service = fs.svc
        ctx.request_context.lifespan_context.drive_folder_cache = MagicMock()
        ctx.report_progress = AsyncMock()
        return ctx

    async def test_default_is_non_recursive_and_ignores_subfolders(self, tmp_path):
        # A subfolder plus export_format set together used to hit the pre-existing bug
        # where a folder's mimeType (starts with "application/vnd.google-apps.", same
        # prefix as real Workspace docs) got treated as an exportable file and failed.
        fs = _FakeDriveFS(
            {
                "root": [
                    _drive_file("readme.txt", "f1"),
                    _drive_folder("sub", "subid"),
                ]
            }
        )
        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            export_format="pdf",
            dry_run=True,
            ctx=self._ctx(fs),
        )
        names = [a["name"] for a in result["actions"]]
        assert names == ["readme.txt"]
        assert result["failed"] == []
        assert result["folders_skipped"] == []
        assert fs.list_calls == ["root"]  # subfolder never queried

    async def test_recursive_both_sides_descends_into_matching_subfolder(self, tmp_path):
        (tmp_path / "sub").mkdir()
        fs = _FakeDriveFS(
            {
                "root": [_drive_folder("sub", "subid")],
                "subid": [_drive_file("nested.txt", "f1")],
            }
        )
        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            dry_run=True,
            recursive=True,
            ctx=self._ctx(fs),
        )
        actions_by_name = {a["name"]: a for a in result["actions"]}
        assert actions_by_name["sub/nested.txt"]["action"] == "download"
        assert result["folders_skipped"] == []
        assert set(fs.list_calls) == {"root", "subid"}

    async def test_recursive_drive_only_subfolder_downloaded_when_direction_allows(self, tmp_path):
        fs = _FakeDriveFS(
            {
                "root": [_drive_folder("sub", "subid")],
                "subid": [_drive_file("nested.txt", "f1")],
            }
        )
        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="bidirectional",
            dry_run=True,
            recursive=True,
            ctx=self._ctx(fs),
        )
        actions_by_name = {a["name"]: a for a in result["actions"]}
        assert actions_by_name["sub/nested.txt"]["action"] == "download"
        assert result["folders_skipped"] == []
        # dry_run never touches the filesystem, even to descend
        assert not (tmp_path / "sub").exists()

    async def test_recursive_drive_only_subfolder_skipped_under_upload_direction(self, tmp_path):
        fs = _FakeDriveFS(
            {
                "root": [_drive_folder("sub", "subid")],
                "subid": [_drive_file("nested.txt", "f1")],
            }
        )
        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="upload",
            dry_run=True,
            recursive=True,
            ctx=self._ctx(fs),
        )
        assert result["folders_skipped"] == ["sub/"]
        assert result["actions"] == []
        assert fs.list_calls == ["root"]  # subfolder's children never fetched

    async def test_recursive_local_only_subfolder_skipped_under_download_direction(self, tmp_path):
        (tmp_path / "sub").mkdir()
        (tmp_path / "sub" / "local.txt").write_text("hi")
        fs = _FakeDriveFS({"root": []})
        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="download",
            dry_run=True,
            recursive=True,
            ctx=self._ctx(fs),
        )
        assert result["folders_skipped"] == ["sub/"]
        assert fs.created_folders == []

    async def test_recursive_local_only_subfolder_uploaded_creates_drive_folder(self, tmp_path):
        (tmp_path / "sub").mkdir()
        (tmp_path / "sub" / "local.txt").write_text("hi")
        fs = _FakeDriveFS({"root": []})
        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="upload",
            recursive=True,
            ctx=self._ctx(fs),
        )
        assert fs.created_folders == [
            {"name": "sub", "mimeType": "application/vnd.google-apps.folder", "parents": ["root"]}
        ]
        assert result["uploaded"] == ["sub/local.txt"]
        assert fs.created_files[0]["name"] == "local.txt"
        assert fs.created_files[0]["parents"] == ["new-folder-0"]
        assert result["folders_skipped"] == []

    async def test_reports_progress_for_each_transfer_not_after_the_whole_batch(self, tmp_path):
        """#316: sync_folder's per-level transfers already run concurrently (#293),
        but nothing reported progress, so a large single-level sync was still silent
        for its whole duration. Progress must be reported inside each transfer's own
        coroutine as it completes — not after the level's asyncio.gather resolves,
        which would only ever deliver one final burst instead of live updates."""
        (tmp_path / "a.txt").write_text("hi")
        (tmp_path / "b.txt").write_text("bye")
        fs = _FakeDriveFS({"root": []})
        ctx = self._ctx(fs)

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="upload",
            ctx=ctx,
        )
        assert set(result["uploaded"]) == {"a.txt", "b.txt"}
        assert ctx.report_progress.await_count == 2
        completed_values = sorted(c.args[0] for c in ctx.report_progress.await_args_list)
        assert completed_values == [1, 2]
        messages = [c.args[2] for c in ctx.report_progress.await_args_list]
        assert any("a.txt" in m for m in messages)
        assert any("b.txt" in m for m in messages)

    async def test_dry_run_reports_no_progress(self, tmp_path):
        """dry_run transfers nothing, so no progress update should fire either."""
        (tmp_path / "a.txt").write_text("hi")
        fs = _FakeDriveFS({"root": []})
        ctx = self._ctx(fs)

        await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="upload",
            dry_run=True,
            ctx=ctx,
        )
        ctx.report_progress.assert_not_awaited()

    async def test_report_progress_failure_does_not_demote_a_successful_upload(self, tmp_path):
        """PR #351 review: ctx.report_progress raising (e.g. a dropped session)
        must not overwrite an already-successful upload's result, and must not
        skip the drive_folder_cache invalidation that a real mutation earns."""
        (tmp_path / "a.txt").write_text("hi")
        fs = _FakeDriveFS({"root": []})
        ctx = self._ctx(fs)
        ctx.report_progress.side_effect = RuntimeError("connection dropped")

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="upload",
            ctx=ctx,
        )
        assert result["uploaded"] == ["a.txt"]
        assert result["failed"] == []
        ctx.request_context.lifespan_context.drive_folder_cache.mark_dirty.assert_called_once_with(
            "root"
        )

    async def test_file_folder_name_collision_recorded_as_failed_not_crashed(self, tmp_path):
        """PR #328 review: a Drive file and a Drive folder can share a name (keyed by
        ID, not name). A local file already occupying the subfolder's target path
        (e.g. downloaded moments earlier at the file level) used to crash the whole
        sync via an uncaught FileExistsError from mkdir(exist_ok=True) — exist_ok only
        tolerates an existing *directory*, not an existing file at the same path."""
        (tmp_path / "collide").write_text("existing local file, not a directory")
        fs = _FakeDriveFS(
            {
                "root": [_drive_folder("collide", "collide-folder-id")],
                "collide-folder-id": [],
            }
        )
        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="download",
            recursive=True,
            ctx=self._ctx(fs),
        )
        assert result["failed"] == [
            {
                "name": "collide/",
                "error": (
                    "cannot create local folder 'collide': a file with the same name "
                    "already exists at this path"
                ),
            }
        ]
        # the pre-existing local file itself must survive untouched
        assert (tmp_path / "collide").read_text() == "existing local file, not a directory"

    async def test_drive_folder_create_failure_recorded_as_failed_not_crashed(self, tmp_path):
        """PR #328 review: unlike every file-level transfer, the Drive folder-create
        call for a local-only subfolder being uploaded had no try/except — a
        transient API error there used to propagate uncaught and abort the entire
        multi-level sync instead of recording one failed item."""
        (tmp_path / "sub").mkdir()
        (tmp_path / "sub" / "local.txt").write_text("hi")

        def _create(**kwargs):
            if kwargs["body"].get("mimeType") == "application/vnd.google-apps.folder":
                resp = MagicMock(status=500)
                raise HttpError(resp=resp, content=b'{"error": {"message": "boom"}}')
            resp = MagicMock()
            resp.execute.return_value = {"id": "new-file"}
            return resp

        svc = MagicMock()
        svc.files.return_value.list.return_value.execute.return_value = {"files": []}
        svc.files.return_value.create.side_effect = _create
        ctx = MagicMock()
        ctx.request_context.lifespan_context.drive_service = svc
        ctx.request_context.lifespan_context.drive_folder_cache = MagicMock()
        ctx.report_progress = AsyncMock()

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="upload",
            recursive=True,
            ctx=ctx,
        )
        assert len(result["failed"]) == 1
        assert result["failed"][0]["name"] == "sub/"
        assert "boom" in result["failed"][0]["error"]
        # the call itself must not raise — proven by reaching this line at all

    async def test_recursive_sibling_subfolders_descend_concurrently(self, tmp_path):
        """PR #328 review: sibling subfolder recursion was awaited one at a time
        instead of gathered, so wall-clock time scaled with the sum of subfolder
        round-trips instead of the max. A synchronization barrier proves both
        siblings' Drive list calls are genuinely in flight together at the same
        time, in real OS threads via execute_in_thread — if recursion regresses to
        sequential awaits, only one call is ever in flight and the barrier times
        out. The barrier must live inside `.execute()`, not `.list()`: `.list()` is
        evaluated eagerly on the event-loop thread while building the call chain,
        before execute_in_thread ever hands off to a worker thread — blocking there
        would freeze the single-threaded event loop itself rather than proving
        cross-thread concurrency."""
        barrier = threading.Barrier(2, timeout=2)

        class _ConcurrentFakeDriveFS(_FakeDriveFS):
            def _list(self, **kwargs):
                folder_id = kwargs["q"].split("'")[1]
                self.list_calls.append(folder_id)
                resp = MagicMock()
                if folder_id in ("alpha-id", "beta-id"):

                    def _execute(*args, folder_id=folder_id, **kwargs):
                        barrier.wait()
                        return {"files": self.children.get(folder_id, [])}

                    resp.execute.side_effect = _execute
                else:
                    resp.execute.return_value = {"files": self.children.get(folder_id, [])}
                return resp

        fs = _ConcurrentFakeDriveFS(
            {
                "root": [_drive_folder("alpha", "alpha-id"), _drive_folder("beta", "beta-id")],
                "alpha-id": [_drive_file("a.txt", "fa")],
                "beta-id": [_drive_file("b.txt", "fb")],
            }
        )
        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            dry_run=True,
            recursive=True,
            ctx=self._ctx(fs),
        )
        names = {a["name"] for a in result["actions"]}
        assert names == {"alpha/a.txt", "beta/b.txt"}
        assert result["failed"] == []


class TestSyncFolderConvertMarkdown:
    """Issue #211: sync_folder's convert_markdown param uploads local .md files via
    Drive's native import conversion (same mechanism as upload_local_file's convert
    param, #188), landing as Google Docs that keep their '.md' name so later syncs
    match them back to their local counterpart instead of re-uploading a duplicate
    every run."""

    def _ctx(self, fs: _FakeDriveFS):
        ctx = MagicMock()
        ctx.request_context.lifespan_context.drive_service = fs.svc
        ctx.request_context.lifespan_context.drive_folder_cache = MagicMock()
        ctx.report_progress = AsyncMock()
        return ctx

    async def test_local_only_md_file_converts_to_google_doc(self, tmp_path):
        (tmp_path / "notes.md").write_text("# Heading\n\nBody text.")
        fs = _FakeDriveFS({"root": []})

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="upload",
            convert_markdown=True,
            ctx=self._ctx(fs),
        )

        assert result["uploaded"] == ["notes.md"]
        assert fs.created_files[0]["name"] == "notes.md"
        assert fs.created_files[0]["mimeType"] == "application/vnd.google-apps.document"
        assert fs.created_files[0]["properties"] == {
            transfer_module._CONVERT_MARKDOWN_SOURCE_PROP: "notes.md"
        }
        create_kwargs = fs.svc.files.return_value.create.call_args_list[-1].kwargs
        assert create_kwargs["media_body"].mimetype() == "text/markdown"
        # TC-D218 root cause: Drive's native import-conversion overwrites the
        # modifiedTime requested on create() with its own "now" — a metadata-only
        # follow-up update() re-stamps it correctly.
        assert len(fs.updated_files) == 1
        assert fs.updated_files[0]["fileId"] == "new-file-1"
        assert "modifiedTime" in fs.updated_files[0]["body"]
        assert "media_body" not in fs.updated_files[0]

    async def test_convert_markdown_false_default_uploads_md_as_plain_text(self, tmp_path):
        (tmp_path / "notes.md").write_text("# Heading")
        fs = _FakeDriveFS({"root": []})

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="upload",
            ctx=self._ctx(fs),
        )

        assert result["uploaded"] == ["notes.md"]
        assert "mimeType" not in fs.created_files[0]

    async def test_resync_matches_existing_converted_doc_by_name_not_reuploaded(self, tmp_path):
        """The converted Doc keeps the '.md' name in Drive, so a resync must match
        it directly against the local file (bypassing the export_format-suffix
        scheme used for other Workspace files) — otherwise it looks 'local only'
        again and gets re-uploaded as a duplicate on every sync."""
        local_file = tmp_path / "notes.md"
        local_file.write_text("# Heading")
        drive_mtime = "2024-06-01T12:00:00.000Z"
        os.utime(
            local_file,
            (
                datetime.fromisoformat(drive_mtime.replace("Z", "+00:00")).timestamp(),
                datetime.fromisoformat(drive_mtime.replace("Z", "+00:00")).timestamp(),
            ),
        )
        fs = _FakeDriveFS(
            {
                "root": [
                    _drive_file(
                        "notes.md",
                        "fa",
                        mtime=drive_mtime,
                        mime="application/vnd.google-apps.document",
                        properties={transfer_module._CONVERT_MARKDOWN_SOURCE_PROP: "notes.md"},
                    )
                ]
            }
        )

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="bidirectional",
            convert_markdown=True,
            ctx=self._ctx(fs),
        )

        assert result["skipped"] == ["notes.md"]
        assert result["uploaded"] == []
        assert fs.created_files == []

    async def test_local_edit_after_conversion_updates_in_place_not_recreated(self, tmp_path):
        local_file = tmp_path / "notes.md"
        local_file.write_text("# Updated heading")
        fs = _FakeDriveFS(
            {
                "root": [
                    _drive_file(
                        "notes.md",
                        "fa",
                        mtime="2020-01-01T00:00:00.000Z",  # far older than local
                        mime="application/vnd.google-apps.document",
                        properties={transfer_module._CONVERT_MARKDOWN_SOURCE_PROP: "notes.md"},
                    )
                ]
            }
        )

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="bidirectional",
            convert_markdown=True,
            ctx=self._ctx(fs),
        )

        assert result["uploaded"] == ["notes.md"]
        assert fs.created_files == []
        assert len(fs.updated_files) == 1
        assert fs.updated_files[0]["fileId"] == "fa"
        assert fs.updated_files[0]["media_body"].mimetype() == "text/markdown"

    async def test_drive_only_converted_doc_reported_as_conflict_not_downloaded(self, tmp_path):
        """No reverse conversion exists (Google Doc -> markdown). A converted Doc
        with no local counterpart yet ('drive only') would normally trigger a
        download — matching it into drive_map without requiring export_format (so
        resyncs of already-converted files work) means this path must be guarded
        explicitly. #414 QA review (finding #4): this must be caught at plan-
        building time and reported as a clean 'conflict', not queued as a doomed
        'download' that only fails at runtime."""
        fs = _FakeDriveFS(
            {
                "root": [
                    _drive_file(
                        "notes.md",
                        "fa",
                        mtime="2030-01-01T00:00:00.000Z",
                        mime="application/vnd.google-apps.document",
                        properties={transfer_module._CONVERT_MARKDOWN_SOURCE_PROP: "notes.md"},
                    )
                ]
            }
        )

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="bidirectional",
            convert_markdown=True,
            ctx=self._ctx(fs),
        )

        assert result["downloaded"] == []
        assert result["failed"] == []
        assert result["conflicts"] == ["notes.md"]

    async def test_plain_file_and_converted_doc_sharing_name_reported_as_failed(self, tmp_path):
        """#422 finding #1: Drive allows a plain file and a convert_markdown Doc to
        share the same display name, and both compute the same drive_map key —
        whichever was enumerated last used to silently win the slot, making the
        other completely invisible to sync (never uploaded, downloaded, or
        reported anywhere). Now reported as a clean failure instead, with neither
        entry touched."""
        fs = _FakeDriveFS(
            {
                "root": [
                    _drive_file("notes.md", "plain-id", mime="text/markdown"),
                    _drive_file(
                        "notes.md",
                        "doc-id",
                        mime="application/vnd.google-apps.document",
                        properties={transfer_module._CONVERT_MARKDOWN_SOURCE_PROP: "notes.md"},
                    ),
                ]
            }
        )

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="bidirectional",
            convert_markdown=True,
            ctx=self._ctx(fs),
        )

        assert len(result["failed"]) == 1
        assert result["failed"][0]["name"] == "notes.md"
        assert "notes.md" in result["failed"][0]["error"]
        assert result["uploaded"] == []
        assert result["downloaded"] == []
        assert result["skipped"] == []
        assert result["conflicts"] == []
        assert fs.created_files == []
        assert fs.updated_files == []

    async def test_two_plain_files_sharing_name_reported_as_failed(self, tmp_path):
        """#422 finding #2: the collision guard originally only fired when
        _is_converted_md differed between the two colliding entries (a plain
        file vs. a convert_markdown Doc) — leaving a same-type collision (e.g.
        two plain files sharing a display name) silently overwritten, the exact
        same bug class in a case the original fix didn't close."""
        fs = _FakeDriveFS(
            {
                "root": [
                    _drive_file("notes.md", "plain-id-1", mime="text/markdown"),
                    _drive_file("notes.md", "plain-id-2", mime="text/plain"),
                ]
            }
        )

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="bidirectional",
            ctx=self._ctx(fs),
        )

        assert len(result["failed"]) == 1
        assert result["failed"][0]["name"] == "notes.md"
        assert "notes.md" in result["failed"][0]["error"]
        assert result["uploaded"] == []
        assert result["downloaded"] == []
        assert result["skipped"] == []
        assert result["conflicts"] == []

    async def test_collision_with_local_counterpart_reported_not_silently_dropped(self, tmp_path):
        """#422 finding #3: a *local* file whose name collides with a Drive-side
        plain-file/converted-Doc pair used to vanish from every output list
        (uploaded/downloaded/skipped/conflicts/actions) with zero acknowledgment
        — the plan loop's bare `continue` skipped it entirely whenever a local
        file happened to share the colliding name. It must now still be reported
        (as a real 'failed' entry for a non-dry-run call), the same as when
        there's no local counterpart at all."""
        (tmp_path / "notes.md").write_text("local content")
        fs = _FakeDriveFS(
            {
                "root": [
                    _drive_file("notes.md", "plain-id", mime="text/markdown"),
                    _drive_file(
                        "notes.md",
                        "doc-id",
                        mime="application/vnd.google-apps.document",
                        properties={transfer_module._CONVERT_MARKDOWN_SOURCE_PROP: "notes.md"},
                    ),
                ]
            }
        )

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="bidirectional",
            convert_markdown=True,
            ctx=self._ctx(fs),
        )

        assert len(result["failed"]) == 1
        assert result["failed"][0]["name"] == "notes.md"
        assert result["uploaded"] == []
        assert result["skipped"] == []
        assert result["conflicts"] == []
        assert fs.created_files == []
        assert fs.updated_files == []

    async def test_collision_dry_run_reports_conflict_not_failed(self, tmp_path):
        """#422 finding #4: dry_run must never report a 'failed' entry — only a
        preview. A name collision used to be written straight to `failed`
        unconditionally, even during dry_run where nothing is materialized,
        breaking that expectation (the pre-existing folder-collision failure
        path in this same function is explicitly guarded against this)."""
        fs = _FakeDriveFS(
            {
                "root": [
                    _drive_file("notes.md", "plain-id", mime="text/markdown"),
                    _drive_file(
                        "notes.md",
                        "doc-id",
                        mime="application/vnd.google-apps.document",
                        properties={transfer_module._CONVERT_MARKDOWN_SOURCE_PROP: "notes.md"},
                    ),
                ]
            }
        )

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="bidirectional",
            convert_markdown=True,
            dry_run=True,
            ctx=self._ctx(fs),
        )

        # #512: dry_run leaves the flat lists empty — 'actions' (asserted below) is
        # the sole, non-redundant source of truth for a dry_run preview.
        assert result["failed"] == []
        assert result["conflicts"] == []
        assert len(result["actions"]) == 1
        assert result["actions"][0]["name"] == "notes.md"
        assert result["actions"][0]["action"] == "collision"

    async def test_drive_only_converted_doc_skipped_not_conflict_under_upload_direction(
        self, tmp_path
    ):
        """#422 finding #3: the drive-only branch checked _is_converted_md before
        checking direction, so a convert_markdown Doc with no local counterpart
        always reported 'conflict' even under direction='upload' — where an
        ordinary (non-converted) drive-only file correctly reports a plain 'skip',
        since an upload-only caller doesn't care about drive-only content at all."""
        fs = _FakeDriveFS(
            {
                "root": [
                    _drive_file(
                        "notes.md",
                        "fa",
                        mime="application/vnd.google-apps.document",
                        properties={transfer_module._CONVERT_MARKDOWN_SOURCE_PROP: "notes.md"},
                    )
                ]
            }
        )

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="upload",
            convert_markdown=True,
            ctx=self._ctx(fs),
        )

        assert result["skipped"] == ["notes.md"]
        assert result["conflicts"] == []
        assert result["downloaded"] == []
        assert result["failed"] == []

    async def test_unrelated_doc_with_matching_md_name_is_not_treated_as_converted(self, tmp_path):
        """#414 QA review (finding #2): a human-created Google Doc that happens to
        be named 'notes.md' but wasn't produced by convert_markdown (no source-name
        property) must not be matched as this tool's converted twin — that would
        let a same-named newer local file silently overwrite its real content."""
        fs = _FakeDriveFS(
            {
                "root": [
                    _drive_file(
                        "notes.md",
                        "fa",
                        mtime="2020-01-01T00:00:00.000Z",
                        mime="application/vnd.google-apps.document",
                        # No properties — an unrelated pre-existing Doc, not one
                        # this tool created.
                    )
                ]
            }
        )

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="bidirectional",
            convert_markdown=True,
            dry_run=True,
            ctx=self._ctx(fs),
        )

        # No export_format and not a recognized converted twin -> excluded from the
        # plan entirely, same as any other Workspace file with no export_format.
        assert result["actions"] == []

    async def test_promoting_a_previously_plain_upload_fails_cleanly(self, tmp_path):
        """#414 QA review (finding #3): a .md file previously synced with
        convert_markdown=False landed as a plain Drive file (not a Doc). Drive's
        API has no supported way to convert an existing file's type via update() —
        only create() honors import conversion — so flipping convert_markdown=True
        later must not silently leave it unpromoted; it should fail with an
        actionable message instead."""
        local_file = tmp_path / "notes.md"
        local_file.write_text("# Updated heading")
        fs = _FakeDriveFS(
            {
                "root": [
                    _drive_file(
                        "notes.md",
                        "fa",
                        mtime="2020-01-01T00:00:00.000Z",  # far older than local
                        mime="text/markdown",  # plain file, never converted
                    )
                ]
            }
        )

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="bidirectional",
            convert_markdown=True,
            ctx=self._ctx(fs),
        )

        assert result["uploaded"] == []
        assert len(result["failed"]) == 1
        assert result["failed"][0]["name"] == "notes.md"
        assert "plain file" in result["failed"][0]["error"]
        assert fs.updated_files == []

    async def test_bidirectional_resync_after_initial_convert_stays_in_sync(self, tmp_path):
        """TC-D218 (#414 QA review): Drive's native import-conversion on create()
        overwrites the modifiedTime we request with its own 'now' once conversion
        finishes, unlike update() which honors it correctly. Without a follow-up
        metadata-only update() re-stamping the correct modifiedTime after create(),
        a bidirectional resync with no local changes reads Drive as newer, tries to
        download the (unconvertible) Doc, and the file gets stuck 'failed' forever."""
        local_file = tmp_path / "notes.md"
        local_file.write_text("# Heading")
        local_mtime_str = datetime.fromtimestamp(
            local_file.stat().st_mtime, tz=timezone.utc
        ).strftime("%Y-%m-%dT%H:%M:%S.000Z")
        fs = _FakeDriveFS({"root": []})
        ctx = self._ctx(fs)

        first = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="bidirectional",
            convert_markdown=True,
            ctx=ctx,
        )
        assert first["uploaded"] == ["notes.md"]

        # The create() call requested the local mtime, but Drive's conversion
        # would normally overwrite it with its own "now" — simulate that by NOT
        # trusting fs.created_files[0]["modifiedTime"] and instead asserting the
        # follow-up update() call is what actually carries the correct value.
        assert len(fs.updated_files) == 1
        assert fs.updated_files[0]["fileId"] == "new-file-1"
        assert fs.updated_files[0]["body"] == {"modifiedTime": local_mtime_str}

        # Simulate Drive's real behavior: the created entry now shows Drive's own
        # (later) modifiedTime rather than what create() was asked to set, exactly
        # as if the metadata-only update() above had never landed — proving the
        # resync only stays clean because that follow-up call fixed it.
        fs.children["root"] = [
            _drive_file(
                "notes.md",
                "new-file-1",
                mtime=local_mtime_str,
                mime="application/vnd.google-apps.document",
                properties={transfer_module._CONVERT_MARKDOWN_SOURCE_PROP: "notes.md"},
            )
        ]

        second = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="bidirectional",
            convert_markdown=True,
            ctx=ctx,
        )
        assert second["skipped"] == ["notes.md"]
        assert second["downloaded"] == []
        assert second["failed"] == []
        assert second["conflicts"] == []

    async def test_resync_without_flag_still_matches_and_does_not_duplicate(self, tmp_path):
        """#414 QA review round 3, finding #1: matching used to be gated on this
        call's own convert_markdown flag, not just the Doc's stamped property. A
        resync that simply omitted convert_markdown=True on a folder containing an
        already-converted Doc saw the local .md as 'local only' and silently
        created a second, plain-text duplicate. Matching must recognize the Doc by
        its property regardless of this call's flag."""
        local_file = tmp_path / "notes.md"
        local_file.write_text("# Heading")
        drive_mtime_str = datetime.fromtimestamp(
            local_file.stat().st_mtime, tz=timezone.utc
        ).strftime("%Y-%m-%dT%H:%M:%S.000Z")
        fs = _FakeDriveFS(
            {
                "root": [
                    _drive_file(
                        "notes.md",
                        "fa",
                        mtime=drive_mtime_str,
                        mime="application/vnd.google-apps.document",
                        properties={transfer_module._CONVERT_MARKDOWN_SOURCE_PROP: "notes.md"},
                    )
                ]
            }
        )

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="bidirectional",
            # convert_markdown intentionally omitted (defaults False) — the Doc was
            # converted on a prior call, this one just forgot the flag.
            ctx=self._ctx(fs),
        )

        assert result["skipped"] == ["notes.md"]
        assert result["uploaded"] == []
        assert fs.created_files == []

    async def test_local_edit_without_flag_reimports_in_place_not_duplicated(self, tmp_path):
        """Companion to the above: a local edit to an already-converted Doc must
        still update() in place (correct reimport mimetype) even when this call
        omits convert_markdown, rather than either duplicating or corrupting the
        Doc with a plain-text/octet-stream re-upload."""
        local_file = tmp_path / "notes.md"
        local_file.write_text("# Updated heading")
        fs = _FakeDriveFS(
            {
                "root": [
                    _drive_file(
                        "notes.md",
                        "fa",
                        mtime="2020-01-01T00:00:00.000Z",  # far older than local
                        mime="application/vnd.google-apps.document",
                        properties={transfer_module._CONVERT_MARKDOWN_SOURCE_PROP: "notes.md"},
                    )
                ]
            }
        )

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="bidirectional",
            ctx=self._ctx(fs),
        )

        assert result["uploaded"] == ["notes.md"]
        assert fs.created_files == []
        assert len(fs.updated_files) == 1
        assert fs.updated_files[0]["fileId"] == "fa"
        assert fs.updated_files[0]["media_body"].mimetype() == "text/markdown"

    async def test_upload_local_file_conversion_recognized_by_sync_folder(self, tmp_path):
        """#414 QA review round 3, finding #2: a Doc converted via
        upload_local_file(convert=True) must be recognized by sync_folder's
        convert_markdown matching too — before the fix, only sync_folder's own
        create() call stamped the marker property, so upload_local_file's converted
        Docs were invisible to it and got silently duplicated on the next
        sync_folder run."""
        local_file = tmp_path / "notes.md"
        local_file.write_text("# Heading")
        drive_mtime_str = datetime.fromtimestamp(
            local_file.stat().st_mtime, tz=timezone.utc
        ).strftime("%Y-%m-%dT%H:%M:%S.000Z")
        # Simulates the Doc upload_local_file(convert=True) would have created —
        # same marker property, same mechanism, different call site.
        fs = _FakeDriveFS(
            {
                "root": [
                    _drive_file(
                        "notes.md",
                        "fa",
                        mtime=drive_mtime_str,
                        mime="application/vnd.google-apps.document",
                        properties={transfer_module._CONVERT_MARKDOWN_SOURCE_PROP: "notes.md"},
                    )
                ]
            }
        )

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="bidirectional",
            convert_markdown=True,
            ctx=self._ctx(fs),
        )

        # Recognized as an in-sync match, not a "local only" file to be uploaded as
        # a second, unconverted duplicate.
        assert result["skipped"] == ["notes.md"]
        assert result["uploaded"] == []
        assert fs.created_files == []


class TestSyncFolderDownloadMtimeRoundTrip:
    """Issue #346: a downloaded file's local mtime defaulted to write time ('now'),
    not Drive's modifiedTime — since 'now' is always later than Drive's original
    timestamp, the next sync saw the file as locally newer and re-uploaded it
    (harmless to content, but wasteful, and repeats on every subsequent sync).
    Fixed by setting the local file's mtime to Drive's modifiedTime after a
    successful download, mirroring what the upload branch already does in
    reverse for the Drive side."""

    def _ctx(self, fs: _FakeDriveFS):
        ctx = MagicMock()
        ctx.request_context.lifespan_context.drive_service = fs.svc
        ctx.request_context.lifespan_context.drive_folder_cache = MagicMock()
        ctx.report_progress = AsyncMock()
        return ctx

    def _workspace_fs(self, drive_mtime: str) -> _FakeDriveFS:
        fs = _FakeDriveFS(
            {
                "root": [
                    _drive_file(
                        "a",
                        "fa",
                        mtime=drive_mtime,
                        mime="application/vnd.google-apps.document",
                    )
                ]
            }
        )
        fs.svc.files.return_value.export.return_value.execute.return_value = b"content"
        return fs

    async def test_downloaded_file_mtime_matches_drive_modifiedtime(self, tmp_path):
        drive_mtime = "2024-06-01T12:00:00.000Z"
        fs = self._workspace_fs(drive_mtime)

        await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="bidirectional",
            export_format="pdf",
            ctx=self._ctx(fs),
        )

        local_mtime = datetime.fromtimestamp((tmp_path / "a.pdf").stat().st_mtime, tz=timezone.utc)
        expected = datetime.fromisoformat(drive_mtime.replace("Z", "+00:00"))
        assert abs((local_mtime - expected).total_seconds()) < 1

    async def test_resync_after_download_reports_skipped_not_reuploaded(self, tmp_path):
        drive_mtime = "2024-06-01T12:00:00.000Z"
        fs = self._workspace_fs(drive_mtime)
        ctx = self._ctx(fs)

        first = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="bidirectional",
            export_format="pdf",
            ctx=ctx,
        )
        assert first["downloaded"] == ["a.pdf"]

        second = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="bidirectional",
            export_format="pdf",
            ctx=ctx,
        )
        assert second["skipped"] == ["a.pdf"]
        assert second["downloaded"] == []
        assert second["uploaded"] == []


class TestSyncFolderUseChecksum:
    """Issue #274: mtime alone can't distinguish real content drift from a
    non-content-changing mtime bump (or vice versa) — upload_local_file in
    particular doesn't stamp modifiedTime the way sync_folder's own upload does,
    so identical content re-downloads forever under mtime-only comparison.
    use_checksum=True adds a content check ahead of the mtime comparison for
    names present on both sides."""

    _CONTENT = b"hello world"
    _MD5 = "5eb63bbbe01eeed093cb22bb8f5acdc3"  # md5(_CONTENT)
    _OTHER_MD5 = "0949f7eb1f66dad39d488d5d22531166"  # md5 of different content

    def _ctx(self, fs: _FakeDriveFS):
        ctx = MagicMock()
        ctx.request_context.lifespan_context.drive_service = fs.svc
        ctx.request_context.lifespan_context.drive_folder_cache = MagicMock()
        ctx.report_progress = AsyncMock()
        return ctx

    def _write_local(self, tmp_path, name, content, mtime_str):
        p = tmp_path / name
        p.write_bytes(content)
        dt = datetime.fromisoformat(mtime_str.replace("Z", "+00:00"))
        os.utime(p, (dt.timestamp(), dt.timestamp()))
        return p

    async def test_checksum_match_skips_despite_mtime_far_apart(self, tmp_path):
        # Drive's modifiedTime is far outside the 5s tolerance of the local file's
        # mtime (mirrors upload_local_file's non-stamped modifiedTime), but the
        # content is identical.
        fs = _FakeDriveFS(
            {"root": [_drive_file("a.txt", "fa", mtime="2024-06-01T00:00:00.000Z", md5=self._MD5)]}
        )
        self._write_local(tmp_path, "a.txt", self._CONTENT, "2020-01-01T00:00:00.000Z")

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            use_checksum=True,
            ctx=self._ctx(fs),
        )
        assert result["skipped"] == ["a.txt"]
        assert result["uploaded"] == []
        assert result["downloaded"] == []
        assert result["conflicts"] == []

    async def test_without_use_checksum_same_mtime_gap_is_not_skipped(self, tmp_path):
        # Control: the exact same mismatched-mtime/matching-content setup as above,
        # but use_checksum defaults to False — mtime alone can't tell these apart,
        # so it's treated as a real conflict (drive newer, bidirectional download
        # would apply but let's use upload direction to force a conflict instead).
        fs = _FakeDriveFS(
            {"root": [_drive_file("a.txt", "fa", mtime="2024-06-01T00:00:00.000Z", md5=self._MD5)]}
        )
        self._write_local(tmp_path, "a.txt", self._CONTENT, "2020-01-01T00:00:00.000Z")

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            direction="upload",
            ctx=self._ctx(fs),
        )
        assert result["skipped"] == []
        assert result["conflicts"] == ["a.txt"]

    async def test_checksum_mismatch_falls_back_to_mtime_decision(self, tmp_path):
        # Content genuinely differs and mtimes disagree beyond tolerance (so the
        # checksum comparison actually runs — see the gating tests below). A
        # mismatch doesn't force any outcome by itself — it just falls through to
        # the same mtime-based decision used when use_checksum=False: local is 20s
        # newer here, so 'upload' under the default bidirectional direction.
        fs = _FakeDriveFS(
            {"root": [_drive_file("a.txt", "fa", mtime="2024-06-01T00:00:00.000Z", md5=self._MD5)]}
        )
        self._write_local(tmp_path, "a.txt", b"different content", "2024-06-01T00:00:20.000Z")

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            use_checksum=True,
            ctx=self._ctx(fs),
        )
        assert result["uploaded"] == ["a.txt"]
        assert result["skipped"] == []
        assert result["conflicts"] == []

    async def test_checksum_not_computed_when_mtimes_already_within_tolerance(
        self, tmp_path, monkeypatch
    ):
        # #274 PR #472 review, finding #3: a pair already "in sync" by mtime alone
        # settles there without paying for a hash read that couldn't change the
        # outcome anyway.
        spy = MagicMock(side_effect=transfer_module._local_md5)
        monkeypatch.setattr(transfer_module, "_local_md5", spy)
        fs = _FakeDriveFS(
            {"root": [_drive_file("a.txt", "fa", mtime="2024-06-01T00:00:00.000Z", md5=self._MD5)]}
        )
        self._write_local(tmp_path, "a.txt", self._CONTENT, "2024-06-01T00:00:02.000Z")

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            use_checksum=True,
            ctx=self._ctx(fs),
        )
        assert result["skipped"] == ["a.txt"]
        spy.assert_not_called()

    async def test_checksum_not_computed_during_dry_run(self, tmp_path, monkeypatch):
        # Same finding #3: dry_run is documented as a cheap, no-transfer preview —
        # it must not read every file's full content to hash it.
        spy = MagicMock(side_effect=transfer_module._local_md5)
        monkeypatch.setattr(transfer_module, "_local_md5", spy)
        fs = _FakeDriveFS(
            {"root": [_drive_file("a.txt", "fa", mtime="2024-06-01T00:00:00.000Z", md5=self._MD5)]}
        )
        self._write_local(tmp_path, "a.txt", self._CONTENT, "2020-01-01T00:00:00.000Z")

        await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            use_checksum=True,
            dry_run=True,
            ctx=self._ctx(fs),
        )
        spy.assert_not_called()

    async def test_local_read_failure_reported_as_failed_not_raised(self, tmp_path, monkeypatch):
        # #274 PR #472 review, finding #1: every other per-item operation in
        # _sync_level degrades to a 'failed' entry instead of raising — a file
        # that becomes unreadable between the directory scan and the hash read
        # (deleted, permission-denied, etc.) must behave the same way, not take
        # down the whole sync_folder call.
        def _boom(path):
            raise OSError("permission denied")

        monkeypatch.setattr(transfer_module, "_local_md5", _boom)
        fs = _FakeDriveFS(
            {"root": [_drive_file("a.txt", "fa", mtime="2024-06-01T00:00:00.000Z", md5=self._MD5)]}
        )
        self._write_local(tmp_path, "a.txt", self._CONTENT, "2020-01-01T00:00:00.000Z")

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            use_checksum=True,
            ctx=self._ctx(fs),
        )
        assert result["failed"] == [{"name": "a.txt", "error": "permission denied"}]
        assert result["skipped"] == []
        assert result["uploaded"] == []
        assert result["downloaded"] == []

    async def test_workspace_file_with_no_checksum_falls_back_to_mtime(self, tmp_path):
        # Google Workspace files have no md5Checksum — use_checksum=True must not
        # crash and must behave exactly as use_checksum=False for these.
        fs = _FakeDriveFS(
            {
                "root": [
                    _drive_file(
                        "Notes",
                        "fa",
                        mtime="2024-06-01T00:00:00.000Z",
                        mime="application/vnd.google-apps.document",
                    )
                ]
            }
        )
        fs.svc.files.return_value.export.return_value.execute.return_value = b"content"
        self._write_local(tmp_path, "Notes.pdf", b"stale local copy", "2020-01-01T00:00:00.000Z")

        result = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            export_format="pdf",
            use_checksum=True,
            ctx=self._ctx(fs),
        )
        assert result["downloaded"] == ["Notes.pdf"]

    async def test_local_md5_matches_known_content_hash(self, tmp_path):
        p = tmp_path / "a.txt"
        p.write_bytes(self._CONTENT)
        assert transfer_module._local_md5(p) == self._MD5


class TestDownloadFolder:
    """PR #328 review: download_folder's own Drive listing had the same folder/
    Workspace-file mimeType conflation bug sync_folder's _list_drive_children was
    fixed for — the two tools don't share that helper, so download_folder needed
    its own fix. Subfolders must always be skipped (this tool never descends into
    them), never handed to export()."""

    def _ctx(self, drive_svc):
        ctx = MagicMock()
        ctx.request_context.lifespan_context.drive_service = drive_svc
        ctx.report_progress = AsyncMock()
        return ctx

    async def test_subfolders_always_skipped_not_exported(self, tmp_path):
        svc = MagicMock()
        svc.files.return_value.list.return_value.execute.return_value = {
            "files": [
                {"id": "sub1", "name": "subdir", "mimeType": "application/vnd.google-apps.folder"},
            ]
        }
        result = await _transfer_tools["download_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            export_format="pdf",
            ctx=self._ctx(svc),
        )
        assert result["skipped"] == ["subdir"]
        assert result["downloaded"] == []
        assert result["failed"] == []
        svc.files.return_value.export.assert_not_called()

    async def test_files_download_concurrently(self, tmp_path):
        """#316: download_folder used to loop sequentially, awaiting one transfer at
        a time — 1.04s/file on a real 217-file folder. A synchronization barrier
        proves two exports are genuinely in flight together, in real OS threads via
        execute_in_thread — a regression back to a sequential loop would only ever
        have one call in flight and the barrier would time out."""
        barrier = threading.Barrier(2, timeout=2)

        def _export(**kwargs):
            resp = MagicMock()

            def _execute(*args, **kwargs):
                barrier.wait()
                return b"content"

            resp.execute.side_effect = _execute
            return resp

        svc = MagicMock()
        svc.files.return_value.list.return_value.execute.return_value = {
            "files": [
                {
                    "id": "doc1",
                    "name": "Doc One",
                    "mimeType": "application/vnd.google-apps.document",
                },
                {
                    "id": "doc2",
                    "name": "Doc Two",
                    "mimeType": "application/vnd.google-apps.document",
                },
            ]
        }
        svc.files.return_value.export.side_effect = _export

        result = await _transfer_tools["download_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            export_format="pdf",
            ctx=self._ctx(svc),
        )
        assert set(result["downloaded"]) == {"Doc One.pdf", "Doc Two.pdf"}
        assert result["failed"] == []

    async def test_reports_progress_as_files_complete(self, tmp_path):
        """#316: the 226s call was silent for its entire duration. Each completed
        transfer must fire a notifications/progress update via ctx.report_progress,
        counted against the known total, instead of arriving all at once (or not
        at all) after every download finishes."""
        svc = MagicMock()
        svc.files.return_value.list.return_value.execute.return_value = {
            "files": [
                {
                    "id": "doc1",
                    "name": "Doc One",
                    "mimeType": "application/vnd.google-apps.document",
                },
                {
                    "id": "doc2",
                    "name": "Doc Two",
                    "mimeType": "application/vnd.google-apps.document",
                },
            ]
        }
        svc.files.return_value.export.return_value.execute.return_value = b"content"
        ctx = self._ctx(svc)

        result = await _transfer_tools["download_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            export_format="pdf",
            ctx=ctx,
        )
        assert result["failed"] == []
        assert ctx.report_progress.await_count == 2
        completed_values = sorted(c.args[0] for c in ctx.report_progress.await_args_list)
        assert completed_values == [1, 2]
        for c in ctx.report_progress.await_args_list:
            assert c.args[1] == 2  # total

    async def test_duplicate_drive_filenames_do_not_race_or_double_count(self, tmp_path):
        """PR #351 review, live-reproduced: Drive allows two files with the same
        name (distinct IDs) in one folder; the local filesystem doesn't. The old
        sequential loop was accidentally safe here (each existence check ran only
        after the previous file had fully written) — the concurrent rewrite must
        dedupe by destination path instead of letting two writers race onto the
        same file and double-count size_bytes."""
        svc = MagicMock()
        svc.files.return_value.list.return_value.execute.return_value = {
            "files": [
                {
                    "id": "doc1",
                    "name": "Report",
                    "mimeType": "application/vnd.google-apps.document",
                },
                {
                    "id": "doc2",
                    "name": "Report",
                    "mimeType": "application/vnd.google-apps.document",
                },
            ]
        }
        svc.files.return_value.export.return_value.execute.return_value = b"content"

        result = await _transfer_tools["download_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            export_format="pdf",
            ctx=self._ctx(svc),
        )
        assert result["downloaded"] == ["Report.pdf"]
        assert len(result["failed"]) == 1
        assert result["failed"][0]["name"] == "Report"
        assert "duplicate filename" in result["failed"][0]["error"]
        assert result["size_bytes"] == len(b"content")
        assert (tmp_path / "Report.pdf").read_bytes() == b"content"

    async def test_report_progress_failure_does_not_demote_a_successful_download(self, tmp_path):
        """PR #351 review: ctx.report_progress raising (e.g. a dropped session)
        must not overwrite an already-successful download's result."""
        svc = MagicMock()
        svc.files.return_value.list.return_value.execute.return_value = {
            "files": [
                {
                    "id": "doc1",
                    "name": "Doc One",
                    "mimeType": "application/vnd.google-apps.document",
                },
            ]
        }
        svc.files.return_value.export.return_value.execute.return_value = b"content"
        ctx = self._ctx(svc)
        ctx.report_progress.side_effect = RuntimeError("connection dropped")

        result = await _transfer_tools["download_folder"](
            folder_id="root",
            local_path=str(tmp_path),
            export_format="pdf",
            ctx=ctx,
        )
        assert result["downloaded"] == ["Doc One.pdf"]
        assert result["failed"] == []


class TestDownloadFile:
    """#486: download_file had zero unit test coverage — download_folder (its
    sibling) has TestDownloadFolder above, but the single-file tool itself was
    only ever exercised indirectly, via error-message strings mentioning it by
    name in other tests."""

    def _ctx(self, drive_svc):
        ctx = MagicMock()
        ctx.request_context.lifespan_context.drive_service = drive_svc
        return ctx

    def _metadata(self, name, mime_type):
        return {"name": name, "mimeType": mime_type}

    async def test_google_doc_export(self, tmp_path):
        svc = MagicMock()
        svc.files.return_value.get.return_value.execute.return_value = self._metadata(
            "My Doc", "application/vnd.google-apps.document"
        )
        svc.files.return_value.export.return_value.execute.return_value = b"pdf bytes"

        result = await _transfer_tools["download_file"](
            file_id="doc1",
            local_path=str(tmp_path),
            export_format="pdf",
            ctx=self._ctx(svc),
        )

        svc.files.return_value.export.assert_called_once_with(
            fileId="doc1", mimeType="application/pdf"
        )
        dest = tmp_path / "My Doc.pdf"
        assert result == {
            "local_path": str(dest),
            "name": "My Doc",
            "size_bytes": len(b"pdf bytes"),
        }
        assert dest.read_bytes() == b"pdf bytes"

    async def test_google_sheet_export_as_xlsx(self, tmp_path):
        svc = MagicMock()
        svc.files.return_value.get.return_value.execute.return_value = self._metadata(
            "My Sheet", "application/vnd.google-apps.spreadsheet"
        )
        svc.files.return_value.export.return_value.execute.return_value = b"xlsx bytes"

        result = await _transfer_tools["download_file"](
            file_id="sheet1",
            local_path=str(tmp_path),
            export_format="xlsx",
            ctx=self._ctx(svc),
        )

        svc.files.return_value.export.assert_called_once_with(
            fileId="sheet1",
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        dest = tmp_path / "My Sheet.xlsx"
        assert result["local_path"] == str(dest)
        assert dest.read_bytes() == b"xlsx bytes"

    async def test_explicit_local_path_overrides_drive_filename(self, tmp_path):
        """When local_path names an exact file (not a directory), that name is
        used as-is rather than Drive's own name + a guessed extension — the
        dest.is_dir() branch in download_file is only taken for directories."""
        svc = MagicMock()
        svc.files.return_value.get.return_value.execute.return_value = self._metadata(
            "My Slides", "application/vnd.google-apps.presentation"
        )
        svc.files.return_value.export.return_value.execute.return_value = b"pptx bytes"
        dest = tmp_path / "custom_name.pptx"

        result = await _transfer_tools["download_file"](
            file_id="slides1",
            local_path=str(dest),
            export_format="pptx",
            ctx=self._ctx(svc),
        )

        svc.files.return_value.export.assert_called_once_with(
            fileId="slides1",
            mimeType="application/vnd.openxmlformats-officedocument.presentationml.presentation",
        )
        assert result["local_path"] == str(dest)
        assert dest.read_bytes() == b"pptx bytes"

    async def test_binary_file_download(self, tmp_path, monkeypatch):
        """Non-Google files skip export() entirely and stream via get_media()/
        MediaIoBaseDownload instead — fake the downloader since it otherwise
        drives a real HTTP range-request loop against request.http."""
        svc = MagicMock()
        svc.files.return_value.get.return_value.execute.return_value = self._metadata(
            "photo.png", "image/png"
        )
        content = b"\x89PNG raw bytes"

        class _FakeDownloader:
            def __init__(self, fh, request):
                self._fh = fh

            def next_chunk(self):
                self._fh.write(content)
                return None, True

        monkeypatch.setattr(transfer_module, "MediaIoBaseDownload", _FakeDownloader)

        result = await _transfer_tools["download_file"](
            file_id="bin1",
            local_path=str(tmp_path),
            ctx=self._ctx(svc),
        )

        svc.files.return_value.export.assert_not_called()
        svc.files.return_value.get_media.assert_called_once_with(fileId="bin1")
        dest = tmp_path / "photo.png"
        assert result == {"local_path": str(dest), "name": "photo.png", "size_bytes": len(content)}
        assert dest.read_bytes() == content

    async def test_nonexistent_file_id_propagates_error(self, tmp_path):
        resp = MagicMock()
        resp.status = 404
        svc = MagicMock()
        svc.files.return_value.get.return_value.execute.side_effect = HttpError(
            resp=resp, content=b'{"error": {"message": "File not found: invalidid123xyz"}}'
        )

        with pytest.raises(HttpError):
            await _transfer_tools["download_file"](
                file_id="invalidid123xyz",
                local_path=str(tmp_path),
                ctx=self._ctx(svc),
            )

    async def test_workspace_file_without_export_format_raises_valueerror(self, tmp_path):
        """PR #547 QA round 1: the only place this can ever be verified — the
        matching live case (TC-D105) is a standing SKIP (⚠️ local-filesystem)."""
        svc = MagicMock()
        svc.files.return_value.get.return_value.execute.return_value = self._metadata(
            "My Doc", "application/vnd.google-apps.document"
        )

        with pytest.raises(ValueError, match="export_format is required"):
            await _transfer_tools["download_file"](
                file_id="doc1",
                local_path=str(tmp_path),
                ctx=self._ctx(svc),
            )
        svc.files.return_value.export.assert_not_called()

    async def test_invalid_export_format_raises_valueerror(self, tmp_path):
        """PR #547 QA round 1: same untested-and-unreachable-live rationale as
        the missing-export_format case above."""
        svc = MagicMock()
        svc.files.return_value.get.return_value.execute.return_value = self._metadata(
            "My Doc", "application/vnd.google-apps.document"
        )

        with pytest.raises(ValueError, match="Unknown export_format 'bogus'"):
            await _transfer_tools["download_file"](
                file_id="doc1",
                local_path=str(tmp_path),
                export_format="bogus",
                ctx=self._ctx(svc),
            )
        svc.files.return_value.export.assert_not_called()


class TestSyncFolderResponseSizeCap:
    """PR #328 review: recursive=True removes the previous implicit bound (one
    folder's direct children) on every result list, especially 'actions' during a
    dry run — nothing enforced the shared response-size safety net other capped
    tools use (issue #235/#242)."""

    def _ctx(self, fs: _FakeDriveFS):
        ctx = MagicMock()
        ctx.request_context.lifespan_context.drive_service = fs.svc
        ctx.request_context.lifespan_context.drive_folder_cache = MagicMock()
        ctx.report_progress = AsyncMock()
        return ctx

    async def test_oversized_result_raises(self, tmp_path, monkeypatch):
        monkeypatch.setattr(response_limits, "MAX_TOOL_RESPONSE_CHARS", 10)
        fs = _FakeDriveFS({"root": [_drive_file("readme.txt", "f1")]})
        with pytest.raises(ValueError, match="safety cap"):
            await _transfer_tools["sync_folder"](
                folder_id="root",
                local_path=str(tmp_path),
                dry_run=True,
                ctx=self._ctx(fs),
            )

    async def test_error_points_to_result_local_path_not_local_path(self, tmp_path, monkeypatch):
        # sync_folder's local_path param already means the sync destination — it
        # can't double as a place to dump the oversized response (#512), so the
        # error must point at the dedicated result_local_path param instead.
        monkeypatch.setattr(response_limits, "MAX_TOOL_RESPONSE_CHARS", 10)
        fs = _FakeDriveFS({"root": [_drive_file("readme.txt", "f1")]})
        with pytest.raises(ValueError) as exc_info:
            await _transfer_tools["sync_folder"](
                folder_id="root",
                local_path=str(tmp_path),
                dry_run=True,
                ctx=self._ctx(fs),
            )
        assert "result_local_path" in str(exc_info.value)

    async def test_result_local_path_bypasses_cap_and_writes_to_disk(self, tmp_path, monkeypatch):
        # #512: unlike the sync destination local_path, result_local_path is the
        # offramp for the *response* — passing it must skip the cap entirely and
        # write the full result to disk instead of raising.
        monkeypatch.setattr(response_limits, "MAX_TOOL_RESPONSE_CHARS", 10)
        fs = _FakeDriveFS({"root": [_drive_file("readme.txt", "f1")]})
        out_dir = tmp_path / "out"
        out_dir.mkdir()
        manifest = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(tmp_path / "sync"),
            dry_run=True,
            result_local_path=str(out_dir),
            ctx=self._ctx(fs),
        )
        assert manifest["folder_id"] == "root"
        assert manifest["dry_run"] is True
        written = json.loads(Path(manifest["local_path"]).read_text())
        assert written["actions"][0]["name"] == "readme.txt"

    async def test_result_local_path_equal_to_local_path_raises(self, tmp_path):
        # QA finding, PR #518 review: local_path is scanned as sync input on every
        # call — writing the result manifest there would show up as a new
        # local-only file on the very next sync (and get uploaded on a real run).
        fs = _FakeDriveFS({"root": [_drive_file("readme.txt", "f1")]})
        sync_dir = tmp_path / "sync"
        with pytest.raises(ValueError, match="result_local_path"):
            await _transfer_tools["sync_folder"](
                folder_id="root",
                local_path=str(sync_dir),
                dry_run=True,
                result_local_path=str(sync_dir),
                ctx=self._ctx(fs),
            )

    async def test_result_local_path_inside_local_path_raises(self, tmp_path):
        fs = _FakeDriveFS({"root": [_drive_file("readme.txt", "f1")]})
        sync_dir = tmp_path / "sync"
        with pytest.raises(ValueError, match="result_local_path"):
            await _transfer_tools["sync_folder"](
                folder_id="root",
                local_path=str(sync_dir),
                dry_run=True,
                result_local_path=str(sync_dir / "nested" / "out.json"),
                ctx=self._ctx(fs),
            )

    async def test_result_local_path_outside_local_path_succeeds(self, tmp_path):
        fs = _FakeDriveFS({"root": [_drive_file("readme.txt", "f1")]})
        sync_dir = tmp_path / "sync"
        out_dir = tmp_path / "out"
        out_dir.mkdir()
        manifest = await _transfer_tools["sync_folder"](
            folder_id="root",
            local_path=str(sync_dir),
            dry_run=True,
            result_local_path=str(out_dir),
            ctx=self._ctx(fs),
        )
        assert manifest["folder_id"] == "root"
