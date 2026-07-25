import asyncio
import base64
import io
import logging
import mimetypes
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import markdown as _md
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload, MediaInMemoryUpload, MediaIoBaseDownload
from mcp.server.fastmcp import Context
from mcp.types import ToolAnnotations

from ...auth import execute_in_thread, thread_http
from ..response_limits import enforce_response_size_cap
from . import _SA_QUOTA_ERROR

logger = logging.getLogger(__name__)

_EXPORT_MIME: dict[str, tuple[str, str]] = {
    "pdf": ("application/pdf", ".pdf"),
    "html": ("text/html", ".html"),
    "txt": ("text/plain", ".txt"),
    "docx": ("application/vnd.openxmlformats-officedocument.wordprocessingml.document", ".docx"),
    "odt": ("application/vnd.oasis.opendocument.text", ".odt"),
    "rtf": ("application/rtf", ".rtf"),
    "epub": ("application/epub+zip", ".epub"),
    "csv": ("text/csv", ".csv"),
    "xlsx": ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", ".xlsx"),
    "ods": ("application/vnd.oasis.opendocument.spreadsheet", ".ods"),
    "pptx": ("application/vnd.openxmlformats-officedocument.presentationml.presentation", ".pptx"),
}

_SYSTEM_FILES = {".DS_Store", "Thumbs.db", "desktop.ini", ".localized"}

_SYNC_MTIME_TOLERANCE = 5  # seconds — absorbs clock skew and upload-time drift

# Custom Drive file property set on a Google Doc created via convert_markdown's
# native import conversion, recording the exact local filename it was created
# from. Matching converted-md Docs back to their local file by this property
# (rather than by current Drive display name + mimeType alone) means an
# unrelated pre-existing Doc a human happened to name "notes.md" never matches
# (it lacks the property), and a later Drive-side rename/case-change of the Doc
# doesn't desync the match either, since the stored source name never changes
# (#414 QA review, findings #2 and #7).
_CONVERT_MARKDOWN_SOURCE_PROP = "geeSweetConvertMarkdownSource"

# extension -> (source mimeType to upload as, target Google Workspace mimeType to
# request via Drive's native import conversion). Distinct from _EXPORT_MIME above,
# which maps the other direction (Google type -> downloadable export format).
_CONVERT_MIME: dict[str, tuple[str, str]] = {
    ".csv": ("text/csv", "application/vnd.google-apps.spreadsheet"),
    ".xlsx": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.google-apps.spreadsheet",
    ),
    ".docx": (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.google-apps.document",
    ),
    ".md": ("text/markdown", "application/vnd.google-apps.document"),
    ".html": ("text/html", "application/vnd.google-apps.document"),
    ".htm": ("text/html", "application/vnd.google-apps.document"),
    ".pptx": (
        "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        "application/vnd.google-apps.presentation",
    ),
}


async def _upload_local_file(
    drive_service,
    local_path: str,
    parent_folder_id: str,
    name: str | None = None,
    skip_if_exists: bool = True,
    convert: bool = False,
) -> dict[str, Any]:
    """Upload a local file to a Drive folder. Shared core behind the upload_local_file
    tool and docs/content.py's insert_local_images (imported cross-package the same
    way docs/content.py imports _SA_QUOTA_ERROR from tools/drive/__init__.py).

    convert=True requests Drive's native import conversion (CSV/XLSX -> Sheets,
    DOCX/MD/HTML -> Docs, PPTX -> Slides) by uploading with the source format's
    mimeType while setting the destination file's mimeType to the target Google
    Workspace type — this is distinct from create_doc_from_file, which parses the
    file locally and rebuilds it via Docs API requests instead of Drive's importer."""
    path = Path(local_path)
    if not path.is_file():
        raise ValueError(f"No file found at {local_path!r}")

    file_name = name or path.name

    convert_mime: tuple[str, str] | None = None
    if convert:
        # Derived from the effective destination name, not local_path's suffix —
        # a name= override changes what conversion applies (#188 QA review, PR #410).
        dest_suffix = Path(file_name).suffix.lower()
        convert_mime = _CONVERT_MIME.get(dest_suffix)
        if convert_mime is None:
            supported = ", ".join(sorted(_CONVERT_MIME))
            return {
                "error": (
                    f"Conversion not supported for extension {dest_suffix!r}. "
                    f"Supported extensions: {supported}"
                )
            }

    if skip_if_exists:
        safe_name = file_name.replace("\\", "\\\\").replace("'", "\\'")
        existing = await execute_in_thread(
            drive_service.files()
            .list(
                q=f"name='{safe_name}' and '{parent_folder_id}' in parents and trashed=false",
                spaces="drive",
                includeItemsFromAllDrives=True,
                supportsAllDrives=True,
                fields="files(id, name, webViewLink, mimeType)",
                pageSize=1,
            )
            .execute,
            drive_service,
        )
        hits = existing.get("files", [])
        # When converting, a name-only match isn't good enough — an existing file
        # with the same name but the *raw* (unconverted) mimeType isn't actually
        # the converted duplicate skip_if_exists is meant to detect (#188 QA review,
        # PR #410). Only skip if it's already in the target Workspace format;
        # otherwise fall through and upload/convert normally.
        if hits and (convert_mime is None or hits[0].get("mimeType") == convert_mime[1]):
            logger.debug("Skipping upload — %s already exists as %s", file_name, hits[0]["id"])
            return {
                "fileId": hits[0]["id"],
                "name": hits[0]["name"],
                "web_link": hits[0].get("webViewLink"),
                "skipped": True,
            }

    metadata: dict[str, Any] = {"name": file_name, "parents": [parent_folder_id]}
    if convert_mime is not None:
        mime, target_mime = convert_mime
        metadata["mimeType"] = target_mime
    else:
        mime, _ = mimetypes.guess_type(local_path)
        mime = mime or "application/octet-stream"
    media = MediaFileUpload(local_path, mimetype=mime, resumable=True)
    try:
        result = await execute_in_thread(
            drive_service.files()
            .create(
                body=metadata,
                media_body=media,
                supportsAllDrives=True,
                fields="id, name, webViewLink",
            )
            .execute,
            drive_service,
        )
    except HttpError as e:
        if e.resp.status == 403 and b"storageQuotaExceeded" in (e.content or b""):
            return {"error": _SA_QUOTA_ERROR}
        raise

    logger.debug("Uploaded %s → %s (%s)", local_path, result.get("id"), mime)
    return {
        "fileId": result.get("id"),
        "name": result.get("name", file_name),
        "web_link": result.get("webViewLink"),
        "skipped": False,
    }


async def _list_drive_children(drive_service, folder_id: str) -> tuple[list[dict], list[dict]]:
    """Return (files, folders) among the direct children of a Drive folder."""
    results = await execute_in_thread(
        drive_service.files()
        .list(
            q=f"'{folder_id}' in parents and trashed=false",
            spaces="drive",
            includeItemsFromAllDrives=True,
            supportsAllDrives=True,
            fields="files(id, name, mimeType, modifiedTime, properties)",
            pageSize=1000,
        )
        .execute,
        drive_service,
    )
    files: list[dict] = []
    folders: list[dict] = []
    for f in results.get("files", []):
        if f["mimeType"] == "application/vnd.google-apps.folder":
            folders.append(f)
        else:
            files.append(f)
    return files, folders


async def _sync_level(
    lc,
    drive_service,
    drive_folder_id: str | None,
    dest_dir: Path,
    rel_prefix: str,
    direction: str,
    export_format: str | None,
    convert_markdown: bool,
    skip_system_files: bool,
    dry_run: bool,
    recursive: bool,
    uploaded: list[str],
    downloaded: list[str],
    skipped: list[str],
    conflicts: list[str],
    failed: list[dict[str, str]],
    actions: list[dict[str, str]],
    folders_skipped: list[str],
    ctx: Context,
    progress_count: list[int],
) -> int:
    """
    Sync the files directly inside one Drive-folder/local-dir pair and, if `recursive`,
    descend into subfolders matched by name. `drive_folder_id=None` simulates a Drive
    folder that doesn't exist yet (used for dry-run planning of a not-yet-created
    upload-direction folder) — no Drive API call is made and drive-side maps stay empty.

    convert_markdown=True (#211) treats a local .md file's upload as a request for
    Drive's native import conversion (same mechanism as upload_local_file's convert
    param, #188): uploaded with mimetype='text/markdown', landing as a Google Doc
    that keeps the original '.md' name. Since the converted file is still named
    '<name>.md' in Drive, it's matched back to its local counterpart directly (not
    via the export_format suffix scheme used for other Workspace files) so re-syncs
    settle into "in sync" via the normal mtime comparison instead of re-uploading
    a duplicate on every run.

    Returns bytes downloaded at this level and below; all other results are appended
    into the shared accumulator lists/dicts passed in from the top-level call.
    """
    drive_files: list[dict] = []
    drive_folders: list[dict] = []
    if drive_folder_id is not None:
        drive_files, drive_folders = await _list_drive_children(drive_service, drive_folder_id)

    drive_map: dict[str, dict] = {}
    for f in drive_files:
        is_workspace = f["mimeType"].startswith("application/vnd.google-apps.")
        is_converted_md = False
        if is_workspace:
            convert_source = (f.get("properties") or {}).get(_CONVERT_MARKDOWN_SOURCE_PROP)
            is_converted_md = (
                convert_markdown
                and f["mimeType"] == _CONVERT_MIME[".md"][1]
                and convert_source is not None
            )
            if is_converted_md:
                # The stored source name, not f["name"] — see _CONVERT_MARKDOWN_SOURCE_PROP.
                assert convert_source is not None
                local_name = convert_source
            elif not export_format:
                continue  # excluded without an export format
            else:
                local_name = f["name"] + _EXPORT_MIME[export_format][1]
        else:
            local_name = f["name"]
        # _is_converted_md travels with the entry so the plan-building and download
        # logic below can gate on it without re-deriving convert_source (#414 QA
        # review, finding #1: this Doc must never be downloaded via export_format,
        # regardless of whether export_format is set).
        drive_map[local_name] = {**f, "_is_converted_md": is_converted_md}

    local_map: dict[str, Path] = {}
    if dest_dir.is_dir():
        for p in dest_dir.iterdir():
            if not p.is_file():
                continue
            if skip_system_files and p.name in _SYSTEM_FILES:
                continue
            local_map[p.name] = p

    def _drive_mtime(entry: dict) -> datetime:
        return datetime.fromisoformat(entry["modifiedTime"].replace("Z", "+00:00"))

    def _local_mtime(p: Path) -> datetime:
        return datetime.fromtimestamp(p.stat().st_mtime, tz=timezone.utc)

    plan: list[dict[str, str]] = []
    for name in sorted(drive_map.keys() | local_map.keys()):
        in_drive = name in drive_map
        in_local = name in local_map

        if in_drive and not in_local:
            if drive_map[name]["_is_converted_md"]:
                # A convert_markdown Doc has no reverse conversion — queuing this as
                # a "download" here (as the pre-#414 code did) would either crash on
                # the runtime guard below or, worse, write export_format's binary
                # export content into a file still named .md if export_format was
                # also set (#414 QA review, findings #1 and #4). Report it plainly
                # up front instead, in both dry_run and a real run.
                plan.append(
                    {
                        "name": name,
                        "action": "conflict",
                        "reason": (
                            "drive-only convert_markdown Doc has no reverse conversion — "
                            "add a matching local .md or remove it in Drive"
                        ),
                    }
                )
            elif direction in ("download", "bidirectional"):
                plan.append({"name": name, "action": "download", "reason": "drive only"})
            else:
                plan.append(
                    {"name": name, "action": "skip", "reason": "drive only, upload direction"}
                )

        elif in_local and not in_drive:
            if direction in ("upload", "bidirectional"):
                plan.append({"name": name, "action": "upload", "reason": "local only"})
            else:
                plan.append(
                    {"name": name, "action": "skip", "reason": "local only, download direction"}
                )

        else:
            dmtime = _drive_mtime(drive_map[name])
            lmtime = _local_mtime(local_map[name])
            diff = (lmtime - dmtime).total_seconds()

            if abs(diff) <= _SYNC_MTIME_TOLERANCE:
                plan.append({"name": name, "action": "skip", "reason": "in sync"})
            elif diff > 0:
                if direction in ("upload", "bidirectional"):
                    plan.append(
                        {"name": name, "action": "upload", "reason": f"local newer by {diff:.0f}s"}
                    )
                else:
                    plan.append(
                        {
                            "name": name,
                            "action": "conflict",
                            "reason": f"local newer by {diff:.0f}s but direction is download",
                        }
                    )
            elif drive_map[name]["_is_converted_md"]:
                # Same reasoning as the drive-only case above: this Doc can't be
                # downloaded regardless of direction. In steady state the create()-
                # time modifiedTime fix below keeps this from firing, but it's
                # possible in principle (e.g. residual clock skew), and reporting
                # it as a clean conflict beats a runtime download_fail.
                plan.append(
                    {
                        "name": name,
                        "action": "conflict",
                        "reason": (
                            f"drive newer by {-diff:.0f}s but convert_markdown Docs have no "
                            "reverse conversion — re-upload the local file to update Drive"
                        ),
                    }
                )
            else:
                if direction in ("download", "bidirectional"):
                    plan.append(
                        {
                            "name": name,
                            "action": "download",
                            "reason": f"drive newer by {-diff:.0f}s",
                        }
                    )
                else:
                    plan.append(
                        {
                            "name": name,
                            "action": "conflict",
                            "reason": f"drive newer by {-diff:.0f}s but direction is upload",
                        }
                    )

    for step in plan:
        actions.append({**step, "name": f"{rel_prefix}{step['name']}"})

    total_bytes = 0

    if dry_run:
        for step in plan:
            rel_name = f"{rel_prefix}{step['name']}"
            if step["action"] == "skip":
                skipped.append(rel_name)
            elif step["action"] == "conflict":
                conflicts.append(rel_name)
            # upload/download actions aren't materialized during dry_run
    else:

        async def _run_one(step: dict[str, str]) -> dict[str, Any]:
            name = step["name"]
            action = step["action"]

            if action == "skip":
                return {"kind": "skip", "name": name}

            if action == "conflict":
                return {"kind": "conflict", "name": name}

            if action == "upload":
                p = local_map[name]
                convert_this = convert_markdown and p.suffix.lower() == ".md"
                convert_mime, convert_target_mime = _CONVERT_MIME[".md"]
                if convert_this:
                    mime = convert_mime
                else:
                    mime, _ = mimetypes.guess_type(str(p))
                    mime = mime or "application/octet-stream"
                lmtime_str = datetime.fromtimestamp(p.stat().st_mtime, tz=timezone.utc).strftime(
                    "%Y-%m-%dT%H:%M:%S.000Z"
                )

                try:
                    media = MediaFileUpload(str(p), mimetype=mime, resumable=True)
                    if name in drive_map:
                        existing = drive_map[name]
                        fid = existing["id"]
                        if convert_this and existing["mimeType"] != convert_target_mime:
                            # This .md was previously synced with convert_markdown=False
                            # and landed as a plain Drive file. Drive's API has no
                            # supported way to convert an existing file's type via
                            # update() — only create() honors import conversion — so
                            # silently re-uploading here would just overwrite the plain
                            # file's raw content without ever promoting it to a Doc
                            # (#414 QA review, finding #3). Surface this explicitly
                            # instead of doing something that looks like it worked.
                            return {
                                "kind": "upload_fail",
                                "name": name,
                                "error": (
                                    f"'{name}' already exists in Drive as a plain file, "
                                    "not a converted Doc — convert_markdown cannot promote "
                                    "an existing file's type; delete it in Drive and "
                                    "re-sync to convert"
                                ),
                            }
                        # No mimeType here: the existing file is already the
                        # Google Doc convert_this implies, re-uploading text/markdown
                        # content re-imports it in place without changing its type.
                        await execute_in_thread(
                            drive_service.files()
                            .update(
                                fileId=fid,
                                body={"modifiedTime": lmtime_str},
                                media_body=media,
                                supportsAllDrives=True,
                                fields="id",
                            )
                            .execute,
                            drive_service,
                        )
                        logger.debug("Synced (update) %s%s → Drive", rel_prefix, name)
                    else:
                        body: dict[str, Any] = {
                            "name": name,
                            "parents": [drive_folder_id],
                            "modifiedTime": lmtime_str,
                        }
                        if convert_this:
                            body["mimeType"] = convert_target_mime
                            body["properties"] = {_CONVERT_MARKDOWN_SOURCE_PROP: name}
                        created = await execute_in_thread(
                            drive_service.files()
                            .create(
                                body=body,
                                media_body=media,
                                supportsAllDrives=True,
                                fields="id",
                            )
                            .execute,
                            drive_service,
                        )
                        if convert_this:
                            # Drive's native import-conversion on create() overwrites
                            # the modifiedTime we just requested with its own "now"
                            # once the conversion finishes (observed ~14.7s later) —
                            # update() doesn't have this problem, so a metadata-only
                            # follow-up call re-stamps it correctly. Without this, a
                            # bidirectional resync with no local changes reads Drive
                            # as newer, tries to download the (unconvertible) Doc, and
                            # the file gets stuck 'failed' forever (#414 QA review,
                            # TC-D218).
                            await execute_in_thread(
                                drive_service.files()
                                .update(
                                    fileId=created["id"],
                                    body={"modifiedTime": lmtime_str},
                                    supportsAllDrives=True,
                                    fields="id",
                                )
                                .execute,
                                drive_service,
                            )
                        logger.debug("Synced (create) %s%s → Drive", rel_prefix, name)
                    return {"kind": "upload_ok", "name": name}
                except Exception as e:
                    return {"kind": "upload_fail", "name": name, "error": str(e)}

            # action == "download"
            entry = drive_map[name]
            fid = entry["id"]
            is_workspace = entry["mimeType"].startswith("application/vnd.google-apps.")
            if is_workspace and entry["_is_converted_md"]:
                # No reverse conversion exists (Google Doc -> markdown), regardless of
                # export_format — exporting one of these via export_format would write
                # e.g. binary PDF/DOCX export content into a file still named '.md'
                # instead of failing cleanly (#414 QA review, finding #1). The plan-
                # building loop above already keeps this action from being reached in
                # the normal case (queues 'conflict' instead of 'download'); this is a
                # defense-in-depth guard for the same invariant.
                return {
                    "kind": "download_fail",
                    "name": name,
                    "error": (
                        "Cannot download a convert_markdown Doc: no reverse conversion "
                        "exists — edit the local .md file and re-sync to update Drive"
                    ),
                }
            if is_workspace and not export_format:
                # Unreachable in practice: a plain Workspace file with no
                # export_format never enters drive_map (see the build loop above),
                # and a convert_markdown twin is already handled above regardless of
                # export_format. Kept as a defensive fallback rather than relying on
                # that invariant never changing — surfaces a clean error instead of
                # the KeyError _EXPORT_MIME[None] would raise if it ever became
                # reachable.
                return {
                    "kind": "download_fail",
                    "name": name,
                    "error": (
                        "Cannot download native Google Doc without export_format "
                        "(convert_markdown has no reverse conversion)"
                    ),
                }
            dest_file = dest_dir / name
            try:
                if is_workspace:
                    target_mime = _EXPORT_MIME[export_format][0]
                    content = await execute_in_thread(
                        drive_service.files().export(fileId=fid, mimeType=target_mime).execute,
                        drive_service,
                    )
                    if not isinstance(content, bytes):
                        content = content.encode("utf-8")
                    await asyncio.to_thread(dest_file.write_bytes, content)
                else:

                    def _download_to_completion(fid=fid, dest_file=dest_file) -> None:
                        request = drive_service.files().get_media(fileId=fid)
                        request.http = thread_http(drive_service)
                        with dest_file.open("wb") as fh:
                            downloader = MediaIoBaseDownload(fh, request)
                            done = False
                            while not done:
                                _, done = downloader.next_chunk()

                    await asyncio.to_thread(_download_to_completion)
                # Mirror what the upload branch above does in reverse: set the
                # local file's mtime to Drive's modifiedTime so the round trip
                # stays within _SYNC_MTIME_TOLERANCE on the next sync. Without
                # this the local mtime is "now" (write time), which is always
                # later than Drive's original timestamp — the next sync sees the
                # file as locally newer and re-uploads it, indefinitely (#346).
                dtime = _drive_mtime(entry)
                os.utime(dest_file, (dtime.timestamp(), dtime.timestamp()))
                size = dest_file.stat().st_size
                logger.debug("Synced (download) Drive → %s%s (%d bytes)", rel_prefix, name, size)
                return {"kind": "download_ok", "name": name, "bytes": size}
            except Exception as e:
                return {"kind": "download_fail", "name": name, "error": str(e)}

        async def _run_one_with_progress(step: dict[str, str]) -> dict[str, Any]:
            result = await _run_one(step)
            # Report per-item, not after the whole gather resolves — the gather
            # blocks until every concurrent transfer at this level finishes, so
            # reporting afterward would deliver a single silent burst instead of
            # live progress during the wait (the actual complaint in #316).
            if result["kind"] in ("upload_ok", "upload_fail", "download_ok", "download_fail"):
                progress_count[0] += 1
                try:
                    await ctx.report_progress(
                        progress_count[0],
                        None,
                        f"{rel_prefix}{result['name']}: {result['kind']}",
                    )
                except Exception:
                    # The transfer already succeeded or failed on its own terms —
                    # a broken notification channel (e.g. a dropped session) must
                    # not overwrite that outcome with a spurious failure. PR #351
                    # review: this was previously unguarded and a report_progress
                    # exception here would propagate out, turning an already-
                    # successful transfer into a "failed" item at the gather below.
                    logger.debug(
                        "report_progress failed for %s%s", rel_prefix, result["name"], exc_info=True
                    )
            return result

        # return_exceptions=True: every failure path inside _run_one is already caught
        # and converted into a *_fail result dict, but this also lets every in-flight
        # transfer finish before surfacing an unexpected exception, instead of
        # orphaning in-flight uploads/downloads. Thread-pool default (~32 workers)
        # means very large plans queue rather than fully parallelizing — timing stays
        # accurate, just with diminishing returns past that.
        raw = await asyncio.gather(
            *(_run_one_with_progress(step) for step in plan), return_exceptions=True
        )

        level_changed = False
        for step, o in zip(plan, raw, strict=True):
            rel_name = f"{rel_prefix}{step['name']}"
            if isinstance(o, BaseException):
                failed.append({"name": rel_name, "error": str(o)})
                continue
            kind = o["kind"]
            if kind == "skip":
                skipped.append(rel_name)
            elif kind == "conflict":
                conflicts.append(rel_name)
            elif kind == "upload_ok":
                uploaded.append(rel_name)
                level_changed = True
            elif kind == "download_ok":
                downloaded.append(rel_name)
                total_bytes += o["bytes"]
                level_changed = True
            else:  # upload_fail / download_fail
                failed.append({"name": rel_name, "error": o["error"]})

        if level_changed:
            lc.drive_folder_cache.mark_dirty(drive_folder_id)

    if recursive:
        drive_folder_map = {f["name"]: f for f in drive_folders}
        local_folder_map: dict[str, Path] = {}
        if dest_dir.is_dir():
            for p in dest_dir.iterdir():
                if not p.is_dir():
                    continue
                if skip_system_files and p.name in _SYSTEM_FILES:
                    continue
                local_folder_map[p.name] = p

        # (child_drive_id, child_dest_dir, child_rel_prefix) for every subfolder that
        # survives the prep step below and is ready to be recursed into.
        child_calls: list[tuple[str | None, Path, str]] = []

        for name in sorted(drive_folder_map.keys() | local_folder_map.keys()):
            in_drive = name in drive_folder_map
            in_local = name in local_folder_map
            child_rel_prefix = f"{rel_prefix}{name}/"
            child_dest_dir = dest_dir / name

            if in_drive and in_local:
                child_drive_id = drive_folder_map[name]["id"]
            elif in_drive:
                if direction not in ("download", "bidirectional"):
                    folders_skipped.append(child_rel_prefix)
                    continue
                child_drive_id = drive_folder_map[name]["id"]
                if not dry_run:
                    # A Drive file and a Drive folder can share a name (they're keyed
                    # by ID, not name) — if the file-level pass above just downloaded
                    # a same-named file here, exist_ok=True won't save us since the
                    # existing path isn't a directory.
                    if child_dest_dir.exists() and not child_dest_dir.is_dir():
                        failed.append(
                            {
                                "name": child_rel_prefix,
                                "error": (
                                    f"cannot create local folder '{name}': a file with "
                                    "the same name already exists at this path"
                                ),
                            }
                        )
                        continue
                    child_dest_dir.mkdir(parents=True, exist_ok=True)
            else:  # local only
                if direction not in ("upload", "bidirectional"):
                    folders_skipped.append(child_rel_prefix)
                    continue
                if dry_run:
                    child_drive_id = None  # simulate: not created yet
                else:
                    try:
                        created = await execute_in_thread(
                            drive_service.files()
                            .create(
                                body={
                                    "name": name,
                                    "mimeType": "application/vnd.google-apps.folder",
                                    "parents": [drive_folder_id],
                                },
                                supportsAllDrives=True,
                                fields="id",
                            )
                            .execute,
                            drive_service,
                        )
                    except Exception as e:
                        failed.append({"name": child_rel_prefix, "error": str(e)})
                        continue
                    child_drive_id = created["id"]
                    lc.drive_folder_cache.mark_dirty(drive_folder_id)

            child_calls.append((child_drive_id, child_dest_dir, child_rel_prefix))

        async def _descend(
            child_drive_id: str | None, child_dest_dir: Path, child_rel_prefix: str
        ) -> int:
            return await _sync_level(
                lc,
                drive_service,
                child_drive_id,
                child_dest_dir,
                child_rel_prefix,
                direction,
                export_format,
                convert_markdown,
                skip_system_files,
                dry_run,
                recursive,
                uploaded,
                downloaded,
                skipped,
                conflicts,
                failed,
                actions,
                folders_skipped,
                ctx,
                progress_count,
            )

        # Sibling subfolders are independent — descend into all of them concurrently
        # instead of awaiting one at a time, same rationale as the file-level gather
        # above. Shared accumulator lists are safe to append into concurrently since
        # asyncio coroutines never actually run in parallel, only interleaved at
        # await points.
        child_results = await asyncio.gather(
            *(_descend(*c) for c in child_calls), return_exceptions=True
        )
        for (_, _, child_rel_prefix), r in zip(child_calls, child_results, strict=True):
            if isinstance(r, BaseException):
                failed.append({"name": child_rel_prefix, "error": str(r)})
            else:
                total_bytes += r

    return total_bytes


def _xlsx_range_values(ws, range_str: str | None) -> list[list]:
    """Return cell values from an openpyxl worksheet for the given A1 range (or all data)."""
    if not range_str:
        return [[c.value for c in row] for row in ws.iter_rows()]
    cells = ws[range_str]
    # ws[range] returns a tuple-of-tuples for a multi-cell range, a tuple of cells
    # for a single row/column slice, or a single Cell for a single address.
    if isinstance(cells, tuple) and cells and isinstance(cells[0], tuple):
        return [[c.value for c in row] for row in cells]
    if isinstance(cells, tuple):
        return [[c.value for c in cells]]
    return [[cells.value]]


def register(tool):
    @tool(annotations=ToolAnnotations(title="Export File", readOnlyHint=True))
    async def export_file(
        file_id: str,
        export_format: str,
        ctx: Context = None,
    ) -> dict[str, Any]:
        """
        Export or download a file from Google Drive.

        **Prefer `download_file` when saving to disk** — this tool returns raw
        base64-encoded bytes for binary formats (xlsx, pdf, docx, etc.) that require
        manual decoding. Use `export_file` only when you need the file content in-memory.

        For Google Workspace files (Docs, Sheets, Slides) the file is converted to the
        requested format. For non-Google files the raw content is downloaded.

        Supported export_format values:
          All Google types:  'pdf', 'html'
          Google Docs:       'txt', 'docx', 'odt', 'rtf', 'epub'
          Google Sheets:     'csv', 'xlsx', 'ods'
          Google Slides:     'pptx'
          Non-Google files:  'raw'

        Args:
            file_id: The Google Drive file ID.
            export_format: One of the format strings above.

        Returns:
            fileId, name, mime_type, format, encoding ('utf-8' or 'base64'), content.
            Text formats (txt, html, csv, rtf) are returned as plain strings; all others
            are base64-encoded bytes. Raises ValueError if the response exceeds a safety
            cap (default 40,000 characters, set MAX_TOOL_RESPONSE_CHARS to change it) —
            base64 encoding inflates raw file size by ~33%, so binary exports hit this
            cap at a much smaller *file* size than text ones. Call download_file instead
            for anything but small files; it writes raw bytes straight to disk with no
            base64/JSON overhead.
        """
        _TEXT_MIME_PREFIXES = ("text/",)

        drive_service = ctx.request_context.lifespan_context.drive_service

        metadata = await execute_in_thread(
            drive_service.files()
            .get(fileId=file_id, fields="id, name, mimeType", supportsAllDrives=True)
            .execute,
            drive_service,
        )
        file_mime = metadata.get("mimeType", "")
        is_google_workspace = file_mime.startswith("application/vnd.google-apps.")

        if export_format == "raw" or not is_google_workspace:

            def _download_to_completion() -> bytes:
                request = drive_service.files().get_media(fileId=file_id)
                request.http = thread_http(drive_service)
                buf = io.BytesIO()
                downloader = MediaIoBaseDownload(buf, request)
                done = False
                while not done:
                    _, done = downloader.next_chunk()
                return buf.getvalue()

            raw_bytes = await asyncio.to_thread(_download_to_completion)
            target_mime = file_mime
            content_bytes = raw_bytes
        else:
            if export_format not in _EXPORT_MIME:
                raise ValueError(
                    f"Unknown export_format '{export_format}'. "
                    f"Valid options: {', '.join(_EXPORT_MIME)}, raw"
                )
            target_mime = _EXPORT_MIME[export_format][0]
            content_bytes = await execute_in_thread(
                drive_service.files().export(fileId=file_id, mimeType=target_mime).execute,
                drive_service,
            )
            if isinstance(content_bytes, str):
                content_bytes = content_bytes.encode("utf-8")

        is_text = any(target_mime.startswith(p) for p in _TEXT_MIME_PREFIXES)
        result = {
            "fileId": file_id,
            "name": metadata["name"],
            "mime_type": target_mime,
            "format": export_format,
            "encoding": "utf-8" if is_text else "base64",
            "content": content_bytes.decode("utf-8", errors="replace")
            if is_text
            else base64.b64encode(content_bytes).decode("ascii"),
        }
        enforce_response_size_cap(
            result,
            tool_name="export_file",
            hint="Base64 encoding inflates raw file size by ~33%. Call download_file "
            "instead to write the file straight to disk without this overhead, or ",
            local_path_available=False,
        )
        return result

    @tool(annotations=ToolAnnotations(title="List Revisions", readOnlyHint=True))
    async def list_revisions(file_id: str, ctx: Context = None) -> list[dict[str, Any]]:
        """
        List available revisions for a Google Drive file (Sheets, Docs, or any file).

        Returns revisions in chronological order with their ID, timestamp, and the
        user who made the change. Use the revision ID with export_revision to read
        cell data from a historical version of a spreadsheet.

        Note: Google Drive retains all revisions for 30 days, then auto-prunes unless
        keepForever is set on the revision.

        Args:
            file_id: The Google Drive file ID.

        Returns:
            List of revisions, each with revisionId, modifiedTime, modifiedBy, keepForever.
        """
        drive_service = ctx.request_context.lifespan_context.drive_service
        result = await execute_in_thread(
            drive_service.revisions()
            .list(
                fileId=file_id,
                fields="revisions(id,modifiedTime,lastModifyingUser/displayName,keepForever)",
            )
            .execute,
            drive_service,
        )
        return [
            {
                "revisionId": r["id"],
                "modifiedTime": r.get("modifiedTime"),
                "modifiedBy": r.get("lastModifyingUser", {}).get("displayName"),
                "keepForever": r.get("keepForever", False),
            }
            for r in result.get("revisions", [])
        ]

    @tool(annotations=ToolAnnotations(title="Export Revision", readOnlyHint=True))
    async def export_revision(
        file_id: str,
        revision_id: str,
        range: str | None = None,
        sheet: str | None = None,
        ctx: Context = None,
    ) -> dict[str, Any]:
        """
        Export a historical revision of a Google Sheets file and return its cell data.

        Downloads the revision as an XLSX file and returns the values for the requested
        sheet and range. Use list_revisions to find the revision_id.

        Typical recovery workflow:
          1. list_revisions → find the revision ID from the timestamp before data was lost
          2. export_revision → read the affected range from that revision
          3. batch_update_cells → write the recovered values back to the current sheet

        Note: each call downloads the full file — for large spreadsheets this may be slow.

        Args:
            file_id:     The Google Drive file ID of the spreadsheet.
            revision_id: The revision ID from list_revisions.
            range:       A1 notation range to return, e.g. "A1:D20". Omit for all data.
            sheet:       Sheet (tab) name. Defaults to the first sheet.

        Returns:
            revisionId, modifiedTime, sheet name, range, and values as a list of rows.
        """
        import openpyxl

        drive_service = ctx.request_context.lifespan_context.drive_service

        revision = await execute_in_thread(
            drive_service.revisions()
            .get(fileId=file_id, revisionId=revision_id, fields="exportLinks,modifiedTime")
            .execute,
            drive_service,
        )

        xlsx_url = revision.get("exportLinks", {}).get(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        if not xlsx_url:
            raise ValueError(
                f"No XLSX export available for revision {revision_id}. "
                "The file may not be a Google Sheets file."
            )

        # thread_http(drive_service) must be called inside the worker thread's closure, not
        # eagerly here on the event-loop thread — see execute_in_thread's docstring in auth.py.
        _, content = await asyncio.to_thread(lambda: thread_http(drive_service).request(xlsx_url))
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

        ws = wb[sheet] if sheet else wb.active
        sheet_name = ws.title
        values = _xlsx_range_values(ws, range)

        wb.close()
        return {
            "revisionId": revision_id,
            "modifiedTime": revision.get("modifiedTime"),
            "sheet": sheet_name,
            "range": range,
            "values": values,
        }

    @tool(annotations=ToolAnnotations(title="Upload File", destructiveHint=True))
    async def upload_file(
        name: str,
        content: str,
        source_format: str = "text",
        folder_id: str | None = None,
        convert_to_doc: bool = False,
        ctx: Context = None,
    ) -> dict[str, Any]:
        """
        Upload a text file to Google Drive, optionally converting it to a Google Doc.

        Args:
            name: File name (include extension, e.g. 'notes.md', 'report.html').
            content: Text content to upload.
            source_format: How to interpret the content. One of:
                           'markdown' — Markdown text; converted to HTML before upload.
                           'html'     — Raw HTML.
                           'text'     — Plain text (default).
            folder_id: Destination folder ID. Defaults to the configured folder or Drive root.
            convert_to_doc: If True, create a Google Doc instead of a raw file.
                            'markdown' and 'html' sources retain heading, list, and link
                            formatting via Drive's HTML import. 'text' uploads as plain text
                            and Drive converts it (no formatting preserved).

        Returns:
            fileId, name, parent folder ID, and webViewLink of the created file.

        Note:
            Requires OAuth or ADC auth. Service accounts cannot upload files to personal
            Drive (no storage quota). Works on Shared Drives regardless of auth method.
            Check server://auth-status for your current auth method.
        """
        lc = ctx.request_context.lifespan_context
        drive_service = lc.drive_service
        target_folder_id = folder_id or lc.folder_id

        if source_format == "markdown":
            html_body = _md.markdown(content, extensions=["extra"])
            upload_content = (
                f"<!DOCTYPE html><html><head><meta charset='utf-8'></head>"
                f"<body>{html_body}</body></html>"
            ).encode("utf-8")
            upload_mime = "text/html"
        elif source_format == "html":
            upload_content = content.encode("utf-8")
            upload_mime = "text/html"
        else:
            upload_content = content.encode("utf-8")
            upload_mime = "text/plain"

        file_body: dict[str, Any] = {"name": name}
        if target_folder_id:
            file_body["parents"] = [target_folder_id]

        if convert_to_doc:
            file_body["mimeType"] = "application/vnd.google-apps.document"

        media = MediaInMemoryUpload(upload_content, mimetype=upload_mime, resumable=False)
        try:
            result = await execute_in_thread(
                drive_service.files()
                .create(
                    body=file_body,
                    media_body=media,
                    supportsAllDrives=True,
                    fields="id, name, parents, webViewLink",
                )
                .execute,
                drive_service,
            )
        except HttpError as e:
            if e.resp.status == 403 and b"storageQuotaExceeded" in (e.content or b""):
                return {"error": _SA_QUOTA_ERROR}
            raise

        file_id = result.get("id")
        parents = result.get("parents", [])
        logger.debug("Uploaded %s as file %s (convert_to_doc=%s)", name, file_id, convert_to_doc)

        if target_folder_id:
            lc.drive_folder_cache.mark_dirty(target_folder_id)

        return {
            "fileId": file_id,
            "name": result.get("name", name),
            "parent": parents[0] if parents else "root",
            "web_link": result.get("webViewLink"),
        }

    @tool(annotations=ToolAnnotations(title="Upload Local File", destructiveHint=True))
    async def upload_local_file(
        local_path: str,
        parent_folder_id: str,
        name: str | None = None,
        skip_if_exists: bool = True,
        convert: bool = False,
        ctx: Context = None,
    ) -> dict[str, Any]:
        """
        Upload a file from the local filesystem to a Google Drive folder.
        Handles binary and text files (images, PDFs, DOCX, XLSX, scripts, etc.).

        Args:
            local_path: Absolute path to the local file to upload.
            parent_folder_id: ID of the destination Drive folder.
            name: Name to give the file in Drive. Defaults to the local filename.
            skip_if_exists: If True (default), skip the upload and return the
                            existing file's metadata if a file with the same name
                            already exists in the destination folder.
            convert: If True, request Drive's native import conversion instead of
                     uploading as-is: .csv/.xlsx -> Google Sheets, .docx/.md/.html/.htm
                     -> Google Docs, .pptx -> Google Slides. Any other extension
                     returns an error. Default False (upload preserving original format).

        Returns:
            fileId, name, webViewLink, and 'skipped' (True if skip_if_exists fired).

        Note:
            Requires OAuth or ADC auth. Service accounts cannot upload files to personal
            Drive (no storage quota). Works on Shared Drives regardless of auth method.
            Check server://auth-status for your current auth method.
        """
        lc = ctx.request_context.lifespan_context
        drive_service = lc.drive_service

        result = await _upload_local_file(
            drive_service, local_path, parent_folder_id, name, skip_if_exists, convert
        )
        if "error" not in result and not result.get("skipped"):
            lc.drive_folder_cache.mark_dirty(parent_folder_id)
        return result

    @tool(annotations=ToolAnnotations(title="Upload Local Folder", destructiveHint=True))
    async def upload_local_folder(
        local_path: str,
        parent_folder_id: str,
        skip_if_exists: bool = True,
        skip_system_files: bool = True,
        ctx: Context = None,
    ) -> dict[str, Any]:
        """
        Upload all files in a local directory (non-recursive) to a Google Drive folder.

        Args:
            local_path: Absolute path to the local directory.
            parent_folder_id: ID of the destination Drive folder.
            skip_if_exists: Skip files that already exist in the destination (default True).
            skip_system_files: Skip OS metadata files like .DS_Store (default True).

        Returns:
            Summary with lists of 'uploaded', 'skipped', and 'failed' filenames.

        Note:
            Requires OAuth or ADC auth. Service accounts cannot upload files to personal
            Drive (no storage quota). Works on Shared Drives regardless of auth method.
            Check server://auth-status for your current auth method.
        """
        lc = ctx.request_context.lifespan_context
        drive_service = lc.drive_service

        folder = Path(local_path)
        if not folder.is_dir():
            raise ValueError(f"No directory found at {local_path!r}")

        candidates = [p for p in folder.iterdir() if p.is_file()]
        if skip_system_files:
            candidates = [p for p in candidates if p.name not in _SYSTEM_FILES]

        uploaded: list[str] = []
        skipped: list[str] = []
        failed: list[dict[str, str]] = []

        if skip_if_exists and candidates:
            existing_resp = await execute_in_thread(
                drive_service.files()
                .list(
                    q=f"'{parent_folder_id}' in parents and trashed=false",
                    spaces="drive",
                    includeItemsFromAllDrives=True,
                    supportsAllDrives=True,
                    fields="files(name)",
                    pageSize=1000,
                )
                .execute,
                drive_service,
            )
            existing_names = {f["name"] for f in existing_resp.get("files", [])}
        else:
            existing_names = set()

        for p in sorted(candidates):
            if skip_if_exists and p.name in existing_names:
                skipped.append(p.name)
                continue

            mime, _ = mimetypes.guess_type(str(p))
            mime = mime or "application/octet-stream"

            try:
                metadata: dict[str, Any] = {"name": p.name, "parents": [parent_folder_id]}
                media = MediaFileUpload(str(p), mimetype=mime, resumable=True)
                await execute_in_thread(
                    drive_service.files()
                    .create(
                        body=metadata,
                        media_body=media,
                        supportsAllDrives=True,
                        fields="id",
                    )
                    .execute,
                    drive_service,
                )
                uploaded.append(p.name)
                logger.debug("Uploaded %s (%s)", p.name, mime)
            except Exception as e:
                failed.append({"name": p.name, "error": str(e)})

        if uploaded:
            lc.drive_folder_cache.mark_dirty(parent_folder_id)

        return {"uploaded": uploaded, "skipped": skipped, "failed": failed}

    @tool(annotations=ToolAnnotations(title="Download File", readOnlyHint=True))
    async def download_file(
        file_id: str,
        local_path: str,
        export_format: str | None = None,
        ctx: Context = None,
    ) -> dict[str, Any]:
        """
        Download a file from Google Drive to the local filesystem.

        For non-Google files the raw content is downloaded. For Google Workspace
        files (Docs, Sheets, Slides) an export_format is required to convert the
        file before download.

        If local_path is a directory, the file is saved inside it using the Drive
        filename (with an extension appended for exported Workspace files).

        Args:
            file_id: The Google Drive file ID.
            local_path: Destination file path or directory on the local filesystem.
            export_format: Required for Google Workspace files. One of:
                           Docs   → 'pdf', 'docx', 'html', 'txt', 'odt', 'rtf', 'epub'
                           Sheets → 'pdf', 'xlsx', 'csv', 'ods'
                           Slides → 'pdf', 'pptx'
                           For non-Google files omit this (raw download).

        Returns:
            local_path where the file was written, file name, and byte size.
        """
        drive_service = ctx.request_context.lifespan_context.drive_service

        metadata = await execute_in_thread(
            drive_service.files()
            .get(fileId=file_id, fields="name, mimeType", supportsAllDrives=True)
            .execute,
            drive_service,
        )
        drive_name = metadata["name"]
        file_mime = metadata.get("mimeType", "")
        is_workspace = file_mime.startswith("application/vnd.google-apps.")

        dest = Path(local_path)
        if dest.is_dir():
            if is_workspace and export_format:
                ext = _EXPORT_MIME[export_format][1] if export_format in _EXPORT_MIME else ""
                dest = dest / (drive_name + ext)
            else:
                dest = dest / drive_name

        dest.parent.mkdir(parents=True, exist_ok=True)

        if is_workspace:
            if not export_format:
                raise ValueError(
                    f"export_format is required for Google Workspace file '{drive_name}'. "
                    f"Valid options: {', '.join(_EXPORT_MIME)}"
                )
            if export_format not in _EXPORT_MIME:
                raise ValueError(
                    f"Unknown export_format '{export_format}'. Valid options: {', '.join(_EXPORT_MIME)}"
                )
            target_mime = _EXPORT_MIME[export_format][0]
            content = await execute_in_thread(
                drive_service.files().export(fileId=file_id, mimeType=target_mime).execute,
                drive_service,
            )
            if not isinstance(content, bytes):
                content = content.encode("utf-8")
            await asyncio.to_thread(dest.write_bytes, content)
        else:

            def _download_to_completion() -> None:
                request = drive_service.files().get_media(fileId=file_id)
                request.http = thread_http(drive_service)
                with dest.open("wb") as fh:
                    downloader = MediaIoBaseDownload(fh, request)
                    done = False
                    while not done:
                        _, done = downloader.next_chunk()

            await asyncio.to_thread(_download_to_completion)

        size = dest.stat().st_size
        logger.debug("Downloaded %s → %s (%d bytes)", file_id, dest, size)
        return {"local_path": str(dest), "name": drive_name, "size_bytes": size}

    @tool(annotations=ToolAnnotations(title="Download Folder", readOnlyHint=True))
    async def download_folder(
        folder_id: str,
        local_path: str,
        export_format: str | None = None,
        mime_type_filter: str | None = None,
        skip_if_exists: bool = True,
        ctx: Context = None,
    ) -> dict[str, Any]:
        """
        Download all files in a Google Drive folder (non-recursive) to a local directory.

        For non-Google files the raw content is downloaded. Google Workspace files
        are skipped unless export_format is provided, in which case they are exported
        to that format. Subfolders are always skipped, regardless of export_format —
        this tool never descends into them.

        Files transfer concurrently rather than one at a time. If the caller supplied
        a progressToken, a `notifications/progress` update is sent as each file
        finishes (e.g. "12/217: report.pdf: ok").

        Args:
            folder_id: The Google Drive folder ID.
            local_path: Local directory to download files into (created if needed).
            export_format: If provided, Google Workspace files are exported to this
                           format (e.g. 'pdf', 'docx'). See download_file for full list.
                           Without this, Workspace files are skipped.
            mime_type_filter: Only download files matching this MIME type.
            skip_if_exists: Skip files that already exist at the destination (default True).

        Returns:
            Summary with lists of 'downloaded', 'skipped', and 'failed' filenames,
            plus total 'size_bytes' downloaded.
        """
        drive_service = ctx.request_context.lifespan_context.drive_service
        dest_dir = Path(local_path)
        dest_dir.mkdir(parents=True, exist_ok=True)

        query = f"'{folder_id}' in parents and trashed=false"
        if mime_type_filter:
            safe = mime_type_filter.replace("'", "\\'")
            query += f" and mimeType='{safe}'"

        results = await execute_in_thread(
            drive_service.files()
            .list(
                q=query,
                spaces="drive",
                includeItemsFromAllDrives=True,
                supportsAllDrives=True,
                fields="files(id, name, mimeType)",
                pageSize=1000,
                orderBy="name",
            )
            .execute,
            drive_service,
        )

        downloaded: list[str] = []
        skipped: list[str] = []
        failed: list[dict[str, str]] = []
        total_bytes = 0

        candidates: list[tuple[str, str, bool, Path]] = []
        claimed_dest: set[str] = set()
        for f in results.get("files", []):
            fid = f["id"]
            fname = f["name"]
            fmime = f.get("mimeType", "")

            if fmime == "application/vnd.google-apps.folder":
                # Non-recursive: subfolders are never descended into or exported.
                skipped.append(fname)
                continue

            is_workspace = fmime.startswith("application/vnd.google-apps.")

            if is_workspace and not export_format:
                skipped.append(fname)
                continue

            if is_workspace:
                if export_format not in _EXPORT_MIME:
                    failed.append(
                        {"name": fname, "error": f"Unknown export_format '{export_format}'"}
                    )
                    continue
                ext = _EXPORT_MIME[export_format][1]
                dest_file = dest_dir / (fname + ext)
            else:
                dest_file = dest_dir / fname

            if skip_if_exists and dest_file.exists():
                skipped.append(dest_file.name)
                continue

            # Drive allows two files with the same name (distinct IDs) in one
            # folder; the local filesystem doesn't. Concurrent transfers below
            # would otherwise race to write the identical path — keep the first
            # candidate and record the rest as failed instead of silently
            # clobbering content or double-counting size_bytes (PR #351 review,
            # live-reproduced against a fixture folder with duplicate names).
            dest_key = str(dest_file)
            if dest_key in claimed_dest:
                failed.append(
                    {
                        "name": fname,
                        "error": (
                            f"duplicate filename: another file named '{dest_file.name}' "
                            "already claims this destination path — only the first is "
                            "downloaded"
                        ),
                    }
                )
                continue
            claimed_dest.add(dest_key)

            candidates.append((fid, fname, is_workspace, dest_file))

        total = len(candidates)
        completed = 0

        async def _download_one(
            fid: str, fname: str, is_workspace: bool, dest_file: Path
        ) -> dict[str, Any]:
            nonlocal completed
            try:
                if is_workspace:
                    target_mime = _EXPORT_MIME[export_format][0]
                    content = await execute_in_thread(
                        drive_service.files().export(fileId=fid, mimeType=target_mime).execute,
                        drive_service,
                    )
                    if not isinstance(content, bytes):
                        content = content.encode("utf-8")
                    await asyncio.to_thread(dest_file.write_bytes, content)
                else:

                    def _download_to_completion(fid=fid, dest_file=dest_file) -> None:
                        request = drive_service.files().get_media(fileId=fid)
                        request.http = thread_http(drive_service)
                        with dest_file.open("wb") as fh:
                            downloader = MediaIoBaseDownload(fh, request)
                            done = False
                            while not done:
                                _, done = downloader.next_chunk()

                    await asyncio.to_thread(_download_to_completion)

                size = dest_file.stat().st_size
                logger.debug("Downloaded %s → %s (%d bytes)", fid, dest_file, size)
                result: dict[str, Any] = {"kind": "ok", "name": dest_file.name, "bytes": size}
            except Exception as e:
                result = {"kind": "fail", "name": fname, "error": str(e)}

            # Reported here, inside the per-item coroutine, so updates stream in as
            # each concurrent download finishes rather than arriving in one burst
            # after asyncio.gather resolves — see #316.
            completed += 1
            try:
                await ctx.report_progress(
                    completed, total, f"{completed}/{total}: {result['name']}: {result['kind']}"
                )
            except Exception:
                # The download already succeeded or failed on its own terms — a
                # broken notification channel must not overwrite that outcome.
                # PR #351 review: this was previously unguarded and a
                # report_progress exception here would propagate out, turning an
                # already-successful download into a "failed" item at the
                # gather below.
                logger.debug("report_progress failed for %s", result["name"], exc_info=True)
            return result

        # Concurrent fan-out, same pattern as _sync_level's _run_one: previously this
        # was a sequential `for` loop awaiting one transfer at a time (#316), which
        # measured 1.04s/file and scaled linearly with folder size.
        raw = await asyncio.gather(*(_download_one(*c) for c in candidates), return_exceptions=True)

        for c, o in zip(candidates, raw, strict=True):
            if isinstance(o, BaseException):
                failed.append({"name": c[1], "error": str(o)})
                continue
            if o["kind"] == "ok":
                downloaded.append(o["name"])
                total_bytes += o["bytes"]
            else:
                failed.append({"name": o["name"], "error": o["error"]})

        return {
            "downloaded": downloaded,
            "skipped": skipped,
            "failed": failed,
            "size_bytes": total_bytes,
        }

    @tool(annotations=ToolAnnotations(title="Sync Folder", destructiveHint=True))
    async def sync_folder(
        folder_id: str,
        local_path: str,
        direction: str = "bidirectional",
        export_format: str | None = None,
        convert_markdown: bool = False,
        skip_system_files: bool = True,
        dry_run: bool = False,
        recursive: bool = False,
        ctx: Context = None,
    ) -> dict[str, Any]:
        """
        Sync files between a Google Drive folder and a local directory.

        ## Sync logic

        Files are matched by name. For Google Workspace files (Docs, Sheets, Slides),
        the export extension is appended to form the local name — e.g. a Doc called
        'Notes' with export_format='docx' matches the local file 'Notes.docx'.
        Workspace files with no export_format are skipped entirely — except a Doc
        produced by convert_markdown, which keeps its '.md' name and matches its
        local .md file directly regardless of export_format (see convert_markdown
        below).

        For each matched name the action is decided as follows:

          Drive only  + direction includes download  → download
          Drive only  + direction is 'upload'        → skip
          Local only  + direction includes upload    → upload
          Local only  + direction is 'download'      → skip
          Both sides, mtimes within 5 s tolerance    → skip (already in sync)
          Both sides, local newer by > 5 s           → upload  (if direction includes upload)
          Both sides, Drive newer by > 5 s           → download (if direction includes download)
          Both sides, conflict (direction mismatch)  → skip, listed under 'conflicts'

        Modified times are compared in UTC. When a file is uploaded, its Drive
        modifiedTime is set to the local file's mtime so future syncs stay accurate.

        ## direction values

          'bidirectional' (default) — newer side wins; Drive-only files are downloaded,
                                      local-only files are uploaded.
          'upload'                  — only push local changes to Drive; Drive-only files
                                      and Drive-newer files are left alone.
          'download'                — only pull Drive changes locally; local-only files
                                      and local-newer files are left alone.

        ## recursive

        By default (recursive=False) only files directly inside `folder_id` /
        `local_path` are considered — subfolders are ignored entirely, on either side.

        When recursive=True, subfolders matched by name (same rules as files) are
        walked to any depth. A subfolder present on only one side is only descended
        into — and created on the missing side — when `direction` would actually
        create it there:
          - a Drive-only subfolder is downloaded (local dir created) when direction
            includes download; left alone under 'upload' direction.
          - a local-only subfolder is uploaded (Drive folder created) when direction
            includes upload; left alone under 'download' direction.
        Subfolders left alone this way are listed under 'folders_skipped' (relative
        path, trailing '/') instead of being silently ignored.

        ## dry_run

        When dry_run=True no files or folders are created/transferred. The response
        includes an 'actions' list showing every file considered at every level
        visited, with the reason.

        ## Progress

        File transfers within each level run concurrently rather than one at a time.
        If the caller supplied a progressToken, a `notifications/progress` update is
        sent as each individual upload/download completes (skips and conflicts don't
        emit updates — they're free, not transfers). No update is sent during dry_run,
        since nothing is transferred.

        Args:
            folder_id: Google Drive folder ID to sync against.
            local_path: Local directory path to sync against (created if needed).
            direction: 'bidirectional', 'upload', or 'download'.
            export_format: Required to include Workspace files in the sync. They are
                           exported/compared using this format (e.g. 'pdf', 'docx', 'csv').
            convert_markdown: If True, local .md files are uploaded via Drive's native
                           import conversion, landing as Google Docs (still named
                           '<name>.md') instead of raw text files — same mechanism as
                           upload_local_file's convert param (#188). The converted Doc
                           is matched back to its local .md file directly on later
                           syncs (independent of export_format), so edits round-trip
                           normally instead of re-uploading a duplicate every run.
                           Matching only applies to Docs this call itself created (via
                           an internal Drive property) — a pre-existing Doc a human
                           happened to name '<name>.md' is never mistaken for one and
                           never has its content overwritten. There is no reverse
                           conversion: if the Doc is edited in Drive, or has no local
                           counterpart at all, that entry is reported under 'conflicts'
                           instead of downloaded. A .md previously synced with
                           convert_markdown=False can't be promoted to a Doc in place —
                           that entry is reported under 'failed' with an explanatory
                           message; delete it in Drive and re-sync to convert it.
            skip_system_files: Skip .DS_Store and similar OS metadata files (default True).
            dry_run: If True, plan the sync but transfer nothing.
            recursive: If True, also sync matching subfolders at any depth (see above).
                       Defaults to False — a single call only covers the top-level folder.

        Returns:
            uploaded, downloaded, skipped, conflicts, failed lists (relative paths —
            just the filename at the top level, 'subdir/name' for nested matches when
            recursive descends), 'folders_skipped' (relative subfolder paths not
            entered — always empty when recursive=False), size_bytes transferred,
            dry_run flag, and — when dry_run=True — an 'actions' list with
            {name, action, reason} for every file considered at every level visited.
        """
        if direction not in ("bidirectional", "upload", "download"):
            raise ValueError(
                f"direction must be 'bidirectional', 'upload', or 'download', got '{direction}'"
            )
        if export_format and export_format not in _EXPORT_MIME:
            raise ValueError(
                f"Unknown export_format '{export_format}'. Valid: {', '.join(_EXPORT_MIME)}"
            )

        lc = ctx.request_context.lifespan_context
        drive_service = lc.drive_service
        dest_dir = Path(local_path)
        dest_dir.mkdir(parents=True, exist_ok=True)

        uploaded: list[str] = []
        downloaded: list[str] = []
        skipped: list[str] = []
        conflicts: list[str] = []
        failed: list[dict[str, str]] = []
        actions: list[dict[str, str]] = []
        folders_skipped: list[str] = []

        total_bytes = await _sync_level(
            lc,
            drive_service,
            folder_id,
            dest_dir,
            "",
            direction,
            export_format,
            convert_markdown,
            skip_system_files,
            dry_run,
            recursive,
            uploaded,
            downloaded,
            skipped,
            conflicts,
            failed,
            actions,
            folders_skipped,
            ctx,
            [0],
        )

        result: dict[str, Any] = {
            "uploaded": uploaded,
            "downloaded": downloaded,
            "skipped": skipped,
            "conflicts": conflicts,
            "failed": failed,
            "folders_skipped": folders_skipped,
            "size_bytes": total_bytes,
            "dry_run": dry_run,
        }
        if dry_run:
            result["actions"] = actions

        # recursive=True removes the previous implicit bound (one folder's direct
        # children) on every list here, especially 'actions' during a dry run — the
        # decision doc's own reproduction case (22 subfolders / ~225 files) is a
        # realistic scale to hit the cap.
        enforce_response_size_cap(
            result,
            tool_name="sync_folder",
            hint="Recursive syncs can produce very large result lists. Narrow "
            "folder_id, direction, or recursive scope, or ",
            local_path_available=False,
        )
        return result
